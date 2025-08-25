# -*- coding: utf-8 -*-
"""
NTRE.py -- Registry Scanner (واجهة تبويبية مُعاد تصميمها - الإصدار 4.3.0)
- إزالة أزرار الطي وجعل البطاقات ظاهرة دائماً
- منع إخفاء شريط الأدوات
- مؤشرات استيراد القواعد بشريط تقدم ورسائل نجاح
- تحسين أداء فحص القواعد ببناء فهارس مسبقة واختبارات سريعة
- فلترة المالك كشرط أساسي فقط دون إضافته كسبب
- اقتصار الأسباب على مطابقة الكلمة/القاعدة
- رسوم تفاعلية: نقر لتطبيق الفلترة وTooltips ونِسَب
"""

import sys, os, re, json, base64, html, traceback
from pathlib import Path
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from typing import List, Dict, Any, Tuple, Optional, Set

# ====== اعتمادات واجهة ======
try:
    from PyQt5.QtWidgets import (
        QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QFrame,
        QPushButton, QLabel, QLineEdit, QGroupBox, QListWidget, QListWidgetItem,
        QComboBox, QSpinBox, QCheckBox, QTableWidget, QTableWidgetItem, QHeaderView,
        QAbstractItemView, QMessageBox, QFileDialog, QStatusBar, QProgressBar, QToolBar,
        QAction, QDialog, QDialogButtonBox, QTreeWidget, QTreeWidgetItem, QTextEdit,
        QSplitter, QMenu, QRadioButton, QTabWidget, QSizePolicy, QToolButton, QSpacerItem,
        QProgressDialog
    )
    from PyQt5.QtCore import Qt, QThread, pyqtSignal, QByteArray, QEvent, QTimer, QSize, QPoint
    from PyQt5.QtGui import QPixmap, QKeySequence, QIcon, QPainter, QColor, QFont, QCursor
except Exception as e:
    raise ImportError("PyQt5 مطلوب: pip install PyQt5") from e

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except Exception:
    openpyxl = None

try:
    import pyqtgraph as pg
except Exception:
    pg = None

# PyYAML اختياري لدعم القواعد
try:
    import yaml
    HAVE_YAML = True
except Exception:
    HAVE_YAML = False

# ====== اعتمادات Windows Registry ======
try:
    import winreg
except Exception as e:
        raise ImportError("winreg غير متاح. هذه الأداة مخصصة لويندوز.") from e

# pywin32 (اختياري) لجلب مالك المفتاح
try:
    import win32api, win32security, win32con
    HAVE_PYWIN32 = True
except Exception:
    HAVE_PYWIN32 = False

# ================= مسارات وملفات تخزين =================
APP_DIR = Path.home() / ".ntre_ui"
APP_DIR.mkdir(exist_ok=True)
LISTS_FILE = APP_DIR / "lists_registry.json"
RULES_FILE = APP_DIR / "rules.json"
CONFIG_FILE = APP_DIR / "config.json"
BACKUP_FILE_DEFAULT = APP_DIR / "backup_ntre.json"

# ================= شعارات/أيقونات Base64 بسيطة =================
SAFE_LOGO_BASE64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAHgAAAB4CAYAAAA5ZDbfAAAABHNCSVQICAgIfAhkiAAAAAlwSFlz"
    "AAAXEgAAFxIBZ5/SUgAAABl0RVh0Q3JlYXRpb24gVGltZQAwOC8xNy8yMDI1w3t3YAAAAbxJREFU"
    "eJzt2jEOgjAQgNGJ//9m4k6dF6Qm2b0bqk4w0Qxw8Wq0wN8R9jcA1hG2p3pXfYh2cGQhZ5c6m8kE"
    "hQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPwKQbH8nqH0k8V1gQO1r0x2n6q0mH5zWm0Kkz2"
    "m1rC1Q0m2a2rF+o5wq9k5H0E8p6Vd0Jj3W1f0g3kQy3mZsW9y0i0M1c2c1mQ6p0T3e0xL8i3b+f"
    "6m4D0S8p1a1E5BvJf6m0n4Xo4Q5fNQe1z2W3Zqg3f0p7E1kqfQ3pM7g5QyPmh0pX0a8m6mZqVn9"
    "q6JfQJ2k1Q9v1zK5p3afl+z6B3kB3p7H4c1nO1H8b0dXlWl2b0f4fNQAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAD4D/wA0u3kYw0zYQ8AAAAASUVORK5CYII="
)

def pixmap_from_base64(b64: str) -> QPixmap:
    pad = len(b64) % 4
    if pad:
        b64 += "=" * (4 - pad)
    data = base64.b64decode(b64)
    pm = QPixmap()
    pm.loadFromData(QByteArray(data))
    return pm

# أيقونات حديثة محاصرة لكل زر باستخدام رسم ديناميكي
def modern_icon(glyph: str, bg="#3a86ff", fg="#ffffff") -> QIcon:
    pm = QPixmap(36, 36)
    pm.fill(Qt.transparent)
    p = QPainter(pm)
    p.setRenderHint(QPainter.Antialiasing)
    color = QColor(bg)
    p.setBrush(color)
    p.setPen(Qt.NoPen)
    p.drawRoundedRect(0, 0, 36, 36, 10, 10)
    p.setPen(QColor(fg))
    f = QFont("Segoe UI Symbol", 18, QFont.DemiBold)
    p.setFont(f)
    p.drawText(pm.rect(), Qt.AlignCenter, glyph)
    p.end()
    return QIcon(pm)

def icon_for_action(name: str) -> QIcon:
    palette = {
        "scan": ("▶", "#22c55e"),
        "stop": ("⏹", "#ef4444"),
        "refresh": ("⟳", "#06b6d4"),
        "clear": ("🧹", "#f59e0b"),
        "export": ("⤓", "#8b5cf6"),
        "settings": ("⚙", "#3b82f6"),
        "exit": ("⎋", "#ef4444"),
        "add": ("＋", "#22c55e"),
        "edit": ("✎", "#06b6d4"),
        "remove": ("－", "#ef4444"),
        "browse": ("🗂", "#6366f1"),
        "file": ("📄", "#8b5cf6"),
        "info": ("ℹ", "#0ea5e9"),
        "backup_up": ("⤴", "#22c55e"),
        "backup_down": ("⤵", "#06b6d4"),
        "copy": ("📋", "#14b8a6"),
        "go": ("➡", "#0ea5e9"),
        "delete": ("🗑", "#ef4444"),
        "value": ("🧾", "#a855f7"),
        "tab_kw": ("🔤", "#3b82f6"),
        "tab_rules": ("📜", "#8b5cf6"),
        "folder": ("🗂", "#f97316"),
        "apply": ("✓", "#22c55e"),
        "filter": ("🔎", "#10b981"),
        "collapse": ("▾", "#7c3aed"),
        "expand": ("▸", "#7c3aed"),
    }
    glyph, bg = palette.get(name, ("❖", "#3a86ff"))
    return modern_icon(glyph, bg=bg, fg="#ffffff")

# ================= الترجمات (عربي/إنجليزي) =================
L_AR = {
    "title": "أداة فحص السجل (Registry) -- واجهة تبويبية",
    "scan": "بدء الفحص",
    "stop": "إيقاف الفحص",
    "refresh": "تحديث",
    "clear": "مسح",
    "export": "تصدير",
    "export_excel": "تصدير Excel",
    "export_html": "تصدير HTML",
    "settings": "الإعدادات",
    "exit": "خروج",
    "inputs": "معايير الفحص",
    "stats": "الإحصاءات والرسوم",
    "reasons_chart": "أسباب الاشتباه",
    "results": "نتائج الفحص",
    "filter_hint": "فلترة النتائج",
    "keys_list": "مفاتيح السجل (Recursive)",
    "keywords": "كلمات الفحص (مطابقة دقيقة - حساسة لحالة الأحرف)",
    "add": "إضافة",
    "edit": "تعديل",
    "remove": "حذف",
    "remove_all": "حذف الكل",
    "confirm": "تأكيد",
    "confirm_delete_all_rules": "هل تريد حذف جميع القواعد؟ لا يمكن التراجع.",
    "confirm_delete_selected": "هل تريد حذف العناصر المحددة؟",
    "confirm_stop_scan": "هل تريد إيقاف الفحص الحالي؟",
    "confirm_clear_results": "هل تريد مسح نتائج السجل الحالية؟",
    "confirm_exit": "هل تريد الخروج من التطبيق؟",
    "confirm_delete_registry_value": "هل أنت متأكد من حذف قيمة الريجستري المحددة؟ لا يمكن التراجع عن هذا الإجراء.",
    "ok": "موافق",
    "cancel": "إلغاء",
    "browse": "تصفح...",
    "rule_enabled": "مفعّلة",
    "rule_name": "اسم القاعدة",
    "rule_level": "المستوى",
    "rule_path": "المسار",
    "rules_title": "قائمة القواعد",
    "rule_import": "استيراد ملف...",
    "rule_import_folder": "استيراد مجلد...",
    "rule_remove": "حذف",
    "need_openpyxl": "ثبّت openpyxl لتفعيل التصدير: pip install openpyxl",
    "need_yaml": "ثبّت PyYAML لتفعيل الفحص بالقواعد: pip install PyYAML",
    "no_filters": "الرجاء تحديد معايير للفحص قبل البدء.",
    "no_rules": "لا توجد قواعد مفعلة للفحص.",
    "progress": "جاري الفحص...",
    "done": "انتهى الفحص: {} مشبوه من {} عنصر.",
    "saved": "تم الحفظ",
    "loaded": "تم التحميل",
    "tbl_headers": [
        "المفتاح","الخاصية","القيمة","الكلمة المطابقة","نوع القيمة","آخر تعديل","المالك","الحالة","القاعدة المطابقة","الأسباب"
    ],
    "state_ok": "Access",
    "state_denied": "Denied",
    "reason_kw": "مطابقة كلمة",
    "reason_age": "عنصر حديث",
    "reason_owner": "مطابقة المالك",
    "reason_type": "مطابقة النوع",
    "reason_rule": "مطابقة قاعدة",
    "scan_modes": "أوضاع الفحص",
    "display_modes": "وضع عرض النتائج",
    "display_all": "عرض جميع السجلات",
    "display_matched": "عرض المتطابقة فقط",
    "settings_title": "الإعدادات",
    "config_lang": "اللغة",
    "config_theme": "الثيم",
    "config_value_type": "نوع قيمة الريجستري",
    "config_vt_all": "الكل",
    "config_vt_string": "String",
    "config_vt_expand": "ExpandString",
    "config_vt_multi": "MultiString",
    "config_vt_dword": "DWORD",
    "config_vt_qword": "QWORD",
    "config_vt_binary": "Binary",
    "config_age": "فلترة حسب العمر (أيام)",
    "config_use_age": "تفعيل فلترة العمر",
    "config_accounts": "حسابات المالك",
    "config_backup": "النسخ الاحتياطي",
    "backup_create": "إنشاء نسخة احتياطية",
    "backup_restore": "استعادة نسخة احتياطية",
    "config_saved": "تم حفظ الإعدادات",
    "no_results_to_export": "لا توجد نتائج للتصدير.",
    "browse_registry_title": "تصفح السجل",
    "rule_details": "تفاصيل القاعدة",
    "add_key_title": "إضافة مفتاح سجل",
    "add_key_hint": "مثال: HKLM\\SOFTWARE\\Microsoft\\Windows",
    "add_kw_title": "إضافة كلمات",
    "add_kw_hint": "لإضافة أكثر من كلمة، افصل بينها بفاصلة (,)",
    "edit_kw_title": "تعديل كلمة",
    "result_details": "تفاصيل النتيجة",
    "report_title": "تقرير فحص سجل النظام",
    "report_time": "تاريخ ووقت الفحص",
    "report_criteria": "المعايير المستخدمة",
    "criteria_keys": "المفاتيح",
    "criteria_keywords": "الكلمات",
    "criteria_value_type": "نوع القيمة",
    "criteria_age": "عمر (أيام)",
    "criteria_accounts": "الحسابات",
    "note_chart": "(ثبّت pyqtgraph لعرض الرسم)",
    "ctx_delete_value": "حذف القيمة",
    "ctx_edit_value": "تعديل القيمة",
    "ctx_go_to_key": "الانتقال إلى المفتاح في المستعرض",
    "ctx_copy_path": "نسخ المسار",
    "ctx_copy_value": "نسخ القيمة",
    "edit_value_title": "تعديل قيمة الريجستري",
    "edit_value_new": "القيمة الجديدة:",
    "action_done": "تم التنفيذ",
    "action_failed": "فشلت العملية",
    "tab_keywords": "فحص بالكلمات",
    "tab_rules": "فحص بالقواعد",
    "keys_list_rules": "مفاتيح السجل (للفحص بالقواعد)",
    "filter_column": "العمود",
    "filter_mode": "نوع المطابقة",
    "filter_text": "النص",
    "filter_apply": "تطبيق",
    "match_partial": "تطابق جزئي",
    "match_exact": "تطابق كامل",
    "match_regex": "Regex",

    # خيارات فلترة المالك
    "owner_all": "الكل",
    "owner_systems": "حسابات النظام",
    "owner_localsystem": "النظام المحلي",
    "owner_users": "المستخدم الحالي",
}
L_EN = {
    "title": "Registry Scanner -- Tabbed UI",
    "scan": "Scan",
    "stop": "Stop",
    "refresh": "Refresh",
    "clear": "Clear",
    "export": "Export",
    "export_excel": "Export Excel",
    "export_html": "Export HTML",
    "settings": "Settings",
    "exit": "Exit",
    "inputs": "Scan Criteria",
    "stats": "Stats & Charts",
    "reasons_chart": "Suspicion reasons",
    "results": "Scan Results",
    "filter_hint": "Filter",
    "keys_list": "Registry keys (Recursive)",
    "keywords": "Keywords (exact, case-sensitive)",
    "add": "Add",
    "edit": "Edit",
    "remove": "Remove",
    "remove_all": "Remove All",
    "confirm": "Confirm",
    "confirm_delete_all_rules": "Delete all rules? This cannot be undone.",
    "confirm_delete_selected": "Delete selected items?",
    "confirm_stop_scan": "Do you want to stop the current scan?",
    "confirm_clear_results": "Clear current results?",
    "confirm_exit": "Exit the application?",
    "confirm_delete_registry_value": "Are you sure you want to delete the selected registry value? This cannot be undone.",
    "ok": "OK",
    "cancel": "Cancel",
    "browse": "Browse...",
    "rule_enabled": "Enabled",
    "rule_name": "Rule name",
    "rule_level": "Level",
    "rule_path": "Path",
    "rules_title": "Rules list",
    "rule_import": "Import file...",
    "rule_import_folder": "Import folder...",
    "rule_remove": "Remove",
    "need_openpyxl": "Install openpyxl to enable export: pip install openpyxl",
    "need_yaml": "Install PyYAML to enable rule scanning: pip install PyYAML",
    "no_filters": "Please set scan criteria before scanning.",
    "no_rules": "No enabled rules to scan.",
    "progress": "Scanning...",
    "done": "Done: {} suspicious of {} items.",
    "saved": "Saved",
    "loaded": "Loaded",
    "tbl_headers": [
        "Key","Property","Value","Matched keyword","Value type","Last modified","Owner","State","Matched rule","Reasons"
    ],
    "state_ok": "Access",
    "state_denied": "Denied",
    "reason_kw": "Keyword match",
    "reason_age": "Recent item",
    "reason_owner": "Owner match",
    "reason_type": "Type match",
    "reason_rule": "Rule match",
    "scan_modes": "Scan Modes",
    "display_modes": "Display Mode",
    "display_all": "Show all records",
    "display_matched": "Show matched only",
    "settings_title": "Settings",
    "config_lang": "Language",
    "config_theme": "Theme",
    "config_value_type": "Registry Value Type",
    "config_vt_all": "All",
    "config_vt_string": "String",
    "config_vt_expand": "ExpandString",
    "config_vt_multi": "MultiString",
    "config_vt_dword": "DWORD",
    "config_vt_qword": "QWORD",
    "config_vt_binary": "Binary",
    "config_age": "Age filter (days)",
    "config_use_age": "Enable age filter",
    "config_accounts": "Owner accounts",
    "config_backup": "Backup",
    "backup_create": "Create backup",
    "backup_restore": "Restore backup",
    "config_saved": "Settings saved",
    "no_results_to_export": "No results to export.",
    "browse_registry_title": "Registry Browser",
    "rule_details": "Rule details",
    "add_key_title": "Add Registry Key",
    "add_key_hint": "Example: HKLM\\SOFTWARE\\Microsoft\\Windows",
    "add_kw_title": "Add Keywords",
    "add_kw_hint": "To add multiple keywords, separate by comma (,)",
    "edit_kw_title": "Edit Keyword",
    "result_details": "Result Details",
    "report_title": "System Registry Scan Report",
    "report_time": "Scan Date & Time",
    "report_criteria": "Criteria Used",
    "criteria_keys": "Keys",
    "criteria_keywords": "Keywords",
    "criteria_value_type": "Value Type",
    "criteria_age": "Age (days)",
    "criteria_accounts": "Accounts",
    "note_chart": "(Install pyqtgraph to see chart)",
    "ctx_delete_value": "Delete value",
    "ctx_edit_value": "Edit value",
    "ctx_go_to_key": "Go to key in browser",
    "ctx_copy_path": "Copy path",
    "ctx_copy_value": "Copy value",
    "edit_value_title": "Edit Registry Value",
    "edit_value_new": "New value:",
    "action_done": "Done",
    "action_failed": "Action failed",
    "tab_keywords": "Keyword Scan",
    "tab_rules": "Rules Scan",
    "keys_list_rules": "Registry keys (for Rules scan)",
    "filter_column": "Column",
    "filter_mode": "Match mode",
    "filter_text": "Text",
    "filter_apply": "Apply",
    "match_partial": "Partial",
    "match_exact": "Exact",
    "match_regex": "Regex",

    # Owner filter options
    "owner_all": "All",
    "owner_systems": "Systems",
    "owner_localsystem": "LocalSystem",
    "owner_users": "Users",
}

LANG = "ar"
def tr(key: str):
    return (L_AR if LANG=="ar" else L_EN).get(key, key)

# ================= أدوات مطابقة دقيقة =================
def split_tokens(raw: List[str]) -> List[str]:
    out = []
    for s in raw:
        for part in str(s).split(','):
            t = part.strip()
            if t:
                out.append(t)
    return out

# تعديل: حساسية حالة الأحرف (Case-Sensitive) في المطابقة الدقيقة
def exact_token_present(text: str, tokens: List[str]) -> Optional[str]:
    if not text or not tokens:
        return None
    for t in tokens:
        pat = r'(?<![0-9A-Za-z_])' + re.escape(t) + r'(?![0-9A-Za-z_])'
        if re.search(pat, text):
            return t
    return None

# ================= أدوات السجل =================
HIVE_NAME_TO_CONST = {
    "HKLM": winreg.HKEY_LOCAL_MACHINE,
    "HKEY_LOCAL_MACHINE": winreg.HKEY_LOCAL_MACHINE,
    "HKCU": winreg.HKEY_CURRENT_USER,
    "HKEY_CURRENT_USER": winreg.HKEY_CURRENT_USER,
    "HKCR": winreg.HKEY_CLASSES_ROOT,
    "HKEY_CLASSES_ROOT": winreg.HKEY_CLASSES_ROOT,
    "HKU": winreg.HKEY_USERS,
    "HKEY_USERS": winreg.HKEY_USERS,
    "HKCC": winreg.HKEY_CURRENT_CONFIG,
    "HKEY_CURRENT_CONFIG": winreg.HKEY_CURRENT_CONFIG,
}

REG_TYPE_TO_NAME = {
    winreg.REG_SZ: "String",
    winreg.REG_EXPAND_SZ: "ExpandString",
    winreg.REG_MULTI_SZ: "MultiString",
    winreg.REG_DWORD: "DWORD",
    getattr(winreg, "REG_QWORD", 11): "QWORD",
    winreg.REG_BINARY: "Binary",
}

NAME_TO_REG_TYPE = {
    "string": winreg.REG_SZ,
    "expandstring": winreg.REG_EXPAND_SZ,
    "multistring": winreg.REG_MULTI_SZ,
    "dword": winreg.REG_DWORD,
    "qword": getattr(winreg, "REG_QWORD", 11),
    "binary": winreg.REG_BINARY,
}

def parse_registry_path(path: str) -> Tuple[Optional[int], str]:
    p = path.strip().replace('/', '\\')
    if p.lower().startswith("computer\\"):
        p = p[len("computer\\"):]
    p = (p.replace("HKLM:\\", "HKLM\\")
           .replace("HKCU:\\", "HKCU\\")
           .replace("HKCR:\\", "HKCR\\")
           .replace("HKU:\\", "HKU\\")
           .replace("HKCC:\\", "HKCC\\"))
    parts = p.split('\\', 1)
    hive_name = parts[0].strip()
    sub = parts[1] if len(parts) > 1 else ""
    hive = HIVE_NAME_TO_CONST.get(hive_name.upper())
    if hive is None:
        return None, p
    return hive, sub

def filetime_to_datetime(ft: int) -> Optional[datetime]:
    try:
        epoch_start = 116444736000000000
        ts_sec = (ft - epoch_start) / 10_000_000
        return datetime.utcfromtimestamp(ts_sec)
    except Exception:
        return None

def reg_value_to_text(val, typ) -> str:
    try:
        if typ == winreg.REG_BINARY:
            b = bytes(val)
            s = b.hex()
            if len(s) > 256:
                s = s[:256] + "..."
            return s
        elif typ in (winreg.REG_MULTI_SZ,):
            return "; ".join(val)
        elif typ in (winreg.REG_SZ, winreg.REG_EXPAND_SZ):
            return str(val)
        elif typ in (winreg.REG_DWORD, getattr(winreg, "REG_QWORD", 11)):
            return str(val)
        else:
            return str(val)
    except Exception:
        return str(val)

def try_get_owner(hive_const: int, subkey: str) -> str:
    if not HAVE_PYWIN32:
        return "N/A"
    try:
        hive_map = {
            winreg.HKEY_LOCAL_MACHINE: win32con.HKEY_LOCAL_MACHINE,
            winreg.HKEY_CURRENT_USER: win32con.HKEY_CURRENT_USER,
            winreg.HKEY_CLASSES_ROOT: win32con.HKEY_CLASSES_ROOT,
            winreg.HKEY_USERS: win32con.HKEY_USERS,
            winreg.HKEY_CURRENT_CONFIG: win32con.HKEY_CURRENT_CONFIG,
        }
        whive = hive_map.get(hive_const)
        if whive is None:
            return "N/A"
        h = win32api.RegOpenKeyEx(whive, subkey, 0, win32con.KEY_READ | win32con.READ_CONTROL)
        sd = win32security.GetSecurityInfo(h, win32security.SE_REGISTRY_KEY, win32security.OWNER_SECURITY_INFORMATION)
        sid = sd.GetSecurityDescriptorOwner()
        name, domain, _ = win32security.LookupAccountSid(None, sid)
        return f"{domain}\\{name}" if domain else name
    except Exception:
        return "N/A"

def reg_type_name(typ: int) -> str:
    return REG_TYPE_TO_NAME.get(typ, str(typ))

# ================ معايير الفحص ================
@dataclass
class Criteria:
    keys: List[str] = field(default_factory=list)
    keywords: List[str] = field(default_factory=list)
    value_type: str = "all"
    use_age: bool = False
    days: int = 0
    # owner_filter شرط أساسي فقط
    owner_filter: str = "all"  # "all" | "systems" | "localsystem" | "users"
    mode_keywords: bool = True
    mode_rules: bool = False
    display_mode: str = "matched"  # "matched" or "all"

# ================ بنية القواعد المبسطة ================
@dataclass
class RuleSpec:
    path: str
    title: str
    level: str
    enabled: bool
    predicates: List[Dict[str, Any]] = field(default_factory=list)

def load_rules_from_filelist(filelist: List[Dict[str, Any]]) -> List[RuleSpec]:
    specs: List[RuleSpec] = []
    if not HAVE_YAML:
        return specs
    for entry in filelist:
        if not entry.get("enabled", True):
            continue
        path = entry.get("path", "")
        title = entry.get("title", "") or Path(path).stem
        level = entry.get("level", "")
        try:
            text = Path(path).read_text(encoding="utf-8", errors="ignore")
            data = yaml.safe_load(text) or {}
        except Exception:
            continue
        det = data.get("detection", {})
        preds: List[Dict[str, Any]] = []
        if isinstance(det, dict):
            for _, v in det.items():
                if isinstance(v, list):
                    for item in v:
                        if isinstance(item, str):
                            if any(sym in item for sym in ["^", ".*", "(", "\\", "$", "+", "?", "|"]):
                                try:
                                    comp = re.compile(item, re.IGNORECASE)
                                    preds.append({"type": "re", "value": item, "compiled": comp})
                                except Exception:
                                    preds.append({"type": "kw", "value": item})
                            else:
                                preds.append({"type": "kw", "value": item})
                elif isinstance(v, str):
                    item = v
                    if any(sym in item for sym in ["^", ".*", "(", "\\", "$", "+", "?", "|"]):
                        try:
                            comp = re.compile(item, re.IGNORECASE)
                            preds.append({"type": "re", "value": item, "compiled": comp})
                        except Exception:
                            preds.append({"type": "kw", "value": item})
                    else:
                        preds.append({"type": "kw", "value": item})
        # إزالة التكرارات
        seen = set()
        uniq_preds: List[Dict[str, Any]] = []
        for p in preds:
            key = ("re", p.get("value")) if p["type"] == "re" else ("kw", (p.get("value","") or ""))
            if key in seen:
                continue
            seen.add(key)
            uniq_preds.append(p)
        specs.append(RuleSpec(path=path, title=title, level=level, enabled=True, predicates=uniq_preds))
    return specs

def evaluate_rule_predicates(name: str, text: str, spec: RuleSpec) -> bool:
    if not spec.predicates:
        return False
    # تحسين: اختبار سريع عبر فهرس داخلي سيُبنى في الخيط (تمت الاستفادة منه هناك)
    for p in spec.predicates:
        if p["type"] == "kw":
            tok = p.get("value","")
            if exact_token_present(name or "", [tok]) or exact_token_present(text or "", [tok]):
                return True
        elif p["type"] == "re":
            comp = p.get("compiled")
            try:
                if comp and (comp.search(name or "") or comp.search(text or "")):
                    return True
            except Exception:
                continue
    return False

# ================ ثابت فلترة المالك ================
SYSTEM_ACCOUNTS = {
    "NT AUTHORITY\\SYSTEM",
    "NT AUTHORITY\\LocalService",
    "NT AUTHORITY\\NetworkService",
}

def current_user_account() -> Optional[str]:
    try:
        user = os.getlogin()
    except Exception:
        user = None
    if not user:
        return None
    try:
        domain = os.environ.get("USERDOMAIN", "") or ""
        if domain:
            return f"{domain}\\{user}"
    except Exception:
        pass
    return user

# ================ خيط الفحص ================
class RegistryScannerThread(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(list, int)
    error = pyqtSignal(str)

    def __init__(self, crit: Criteria, rules: Optional[List[RuleSpec]] = None):
        super().__init__()
        self.crit = crit
        self._stop = False
        self.rules = rules or []
        self._current_user_cached = current_user_account()

        # بناء فهارس/مصححات مسبقة لتسريع الفحص
        self._kw_tokens = [k.strip() for k in split_tokens(crit.keywords)] if crit.mode_keywords else []
        # فهارس القواعد: مجموعة كلمات بسيطة وRegexes مجمّعة
        self._rule_kw_set: Set[str] = set()
        self._rule_regex_list: List[re.Pattern] = []
        for spec in self.rules:
            for p in spec.predicates:
                if p["type"] == "kw":
                    val = p.get("value","").strip()
                    if val:
                        self._rule_kw_set.add(val)
                elif p["type"] == "re" and p.get("compiled"):
                    self._rule_regex_list.append(p["compiled"])

    def stop(self): self._stop = True

    def _want_type(self, typ: int) -> bool:
        t = self.crit.value_type.lower()
        if t == "all": return True
        if t == "string" and typ == winreg.REG_SZ: return True
        if t == "expandstring" and typ == winreg.REG_EXPAND_SZ: return True
        if t == "multistring" and typ == winreg.REG_MULTI_SZ: return True
        if t == "dword" and typ == winreg.REG_DWORD: return True
        if t == "qword" and typ == getattr(winreg, "REG_QWORD", 11): return True
        if t == "binary" and typ == winreg.REG_BINARY: return True
        return False

    def _match_value_keywords(self, name: str, value_str: str, kw_tokens: List[str]) -> Optional[str]:
        hit = exact_token_present(name or "", kw_tokens)
        if hit: return hit
        hit = exact_token_present(value_str or "", kw_tokens)
        return hit

    def _owner_pass(self, owner: str) -> bool:
        """
        فلترة المالك شرط أساسي لتمرير السجل فقط. لا تُعيد سبباً.
        """
        mode = (self.crit.owner_filter or "all").lower()
        if mode == "all":
            return True
        o = owner or ""
        if not o or o == "N/A":
            return False
        if mode == "systems":
            return o in SYSTEM_ACCOUNTS
        if mode == "localsystem":
            return o == "NT AUTHORITY\\SYSTEM"
        if mode == "users":
            cu = self._current_user_cached or ""
            return bool(cu and (o.lower() == cu.lower() or cu.split("\\")[-1].lower() in o.lower()))
        return True

    def _age_is_recent(self, last_mod: Optional[datetime], days: int) -> bool:
        if not last_mod: return False
        try:
            return (datetime.utcnow() - last_mod) <= timedelta(days=days)
        except Exception:
            return False

    def _full_key_path(self, hive_const: int, subkey: str) -> str:
        name = None
        for k, v in HIVE_NAME_TO_CONST.items():
            if v == hive_const and len(k) <= 4:
                name = k
                break
        if not name:
            if hive_const == winreg.HKEY_LOCAL_MACHINE: name = "HKLM"
            elif hive_const == winreg.HKEY_CURRENT_USER: name = "HKCU"
            elif hive_const == winreg.HKEY_CLASSES_ROOT: name = "HKCR"
            elif hive_const == winreg.HKEY_USERS: name = "HKU"
            elif hive_const == winreg.HKEY_CURRENT_CONFIG: name = "HKCC"
            else:
                name = str(hive_const)
        return f"{name}\\{subkey}" if subkey else name

    def _fast_rule_match(self, name: str, vtext: str) -> Optional[str]:
        """
        تحسين: اختبار سريع عبر مجموعات مسبقة:
        - إذا وُجدت أي كلمة من rule_kw_set كمطابقة دقيقة في name/value -> يعتبر مطابقاً ويُترك تحديد العنوان لاحقاً.
        - Regex: تجربة على name/value.
        نُعيد مجرد True/اسم قاعدة لاحقاً عند المرور على specs لتحديد العنوان الأول المطابق.
        """
        # كلمات بسيطة
        if self._rule_kw_set:
            if exact_token_present(name, list(self._rule_kw_set)) or exact_token_present(vtext, list(self._rule_kw_set)):
                return "__kw__"
        # Regex
        for comp in self._rule_regex_list:
            try:
                if comp.search(name) or comp.search(vtext):
                    return "__re__"
            except Exception:
                continue
        return None

    def _scan_key_recursive(self, hive_const: int, subkey: str,
                            kw_tokens: List[str],
                            use_age: bool, days: int,
                            out: List[Dict[str, Any]], counter: List[int]):
        if self._stop: return
        access = winreg.KEY_READ
        tried_flags = [access]
        if hasattr(winreg, "KEY_WOW64_64KEY"):
            tried_flags.append(access | winreg.KEY_WOW64_64KEY)
            tried_flags.append(access | winreg.KEY_WOW64_32KEY)

        opened = None
        last_write = None
        state = tr("state_denied")
        for flg in tried_flags:
            try:
                opened = winreg.OpenKey(hive_const, subkey, 0, flg)
                info = winreg.QueryInfoKey(opened)
                last_write = filetime_to_datetime(info[2]) if len(info) >= 3 else None
                state = tr("state_ok")
                break
            except Exception:
                opened = None
                continue

        if opened is None:
            counter[0] += 1
            if counter[0] % 50 == 0:
                self.progress.emit(counter[0])
            return

        owner = try_get_owner(hive_const, subkey)
        if not self._owner_pass(owner):
            # فلترة المالك شرط أساسي: نتجاهل المفتاح كاملاً
            try:
                winreg.CloseKey(opened)
            except Exception:
                pass
            return

        try:
            idx = 0
            while True:
                if self._stop: break
                try:
                    vname, vdata, vtype = winreg.EnumValue(opened, idx)
                except OSError:
                    break
                except Exception:
                    break
                idx += 1
                counter[0] += 1
                if counter[0] % 200 == 0:
                    self.progress.emit(counter[0])

                if not self._want_type(vtype):
                    continue

                vtext = reg_value_to_text(vdata, vtype)
                reasons = []
                matched_kw = ""
                matched_rule = ""
                matched_any = False

                # أوضاع الفحص
                if self.crit.mode_keywords and kw_tokens:
                    hit_kw = self._match_value_keywords(vname or "", vtext, kw_tokens)
                    if hit_kw:
                        matched_kw = hit_kw
                        reasons.append(tr("reason_kw"))
                        matched_any = True

                if self.crit.mode_rules and self.rules:
                    # اختبار سريع أولاً
                    fast = self._fast_rule_match(vname or "", vtext)
                    if fast:
                        # تحديد أول RuleSpec مطابق لإرجاع عنوان القاعدة
                        for spec in self.rules:
                            try:
                                if evaluate_rule_predicates(vname or "", vtext, spec):
                                    matched_rule = spec.title
                                    reasons.append(f"{tr('reason_rule')}: {spec.title}")
                                    matched_any = True
                                    break
                            except Exception:
                                continue

                # اقتصار الأسباب على كلمة/قاعدة فقط: لا نضيف النوع/العمر/المالك كأسباب
                # الفلاتر الإضافية (العمر) لا تُؤثر على الإدراج إلا إذا كان display_mode==matched
                include = True
                filters_active = any([
                    (self.crit.mode_keywords and kw_tokens),
                    (self.crit.mode_rules and self.rules)
                ])
                if self.crit.display_mode == "matched" and filters_active:
                    include = matched_any

                if include:
                    out.append({
                        "key": self._full_key_path(hive_const, subkey),
                        "value_name": vname,
                        "value_str": vtext,
                        "matched_kw": matched_kw,
                        "value_type": reg_type_name(vtype),
                        "last_mod": last_write.strftime("%Y-%m-%d %H:%M:%S") if last_write else "N/A",
                        "owner": owner,
                        "state": state,
                        "matched_rule": matched_rule,
                        "reasons": reasons,
                        "matched_any": matched_any,
                        "hive_const": hive_const,
                        "subkey": subkey,
                        "value_type_raw": vtype,
                    })
        except Exception:
            pass

        try:
            sub_count = winreg.QueryInfoKey(opened)[0]
        except Exception:
            sub_count = 0

        for i in range(sub_count):
            if self._stop: break
            try:
                name = winreg.EnumKey(opened, i)
            except Exception:
                continue
            child = f"{subkey}\\{name}" if subkey else name
            try:
                self._scan_key_recursive(hive_const, child, kw_tokens, use_age, days, out, counter)
            except RecursionError:
                continue
            except Exception:
                continue

        try:
            winreg.CloseKey(opened)
        except Exception:
            pass

    def run(self):
        try:
            crit = self.crit
            kw_tokens = self._kw_tokens
            use_age = crit.use_age
            days = int(crit.days or 0)

            active = any([
                crit.keys,
                (crit.mode_keywords and kw_tokens),
                (crit.mode_rules and self.rules),
                (use_age and days > 0),
                (crit.value_type.lower() != "all"),
                (crit.owner_filter.lower() != "all"),
            ])
            if not active:
                self.finished.emit([], 0)
                return

            results: List[Dict[str, Any]] = []
            counter = [0]
            for raw in crit.keys:
                if self._stop: break
                if not raw or not str(raw).strip():
                    continue
                hive, sub = parse_registry_path(str(raw))
                if hive is None:
                    continue
                self._scan_key_recursive(hive, sub, kw_tokens, use_age, days, results, counter)

            self.finished.emit(results, counter[0])
        except Exception as e:
            self.error.emit(str(e))

# ================ كارد تجميلي (بدون طي) ================
class Card(QFrame):
    def __init__(self, title: str, collapsible: bool = False):
        super().__init__()
        self.setObjectName("Card")
        self.setFrameShape(QFrame.StyledPanel)
        self.setFrameShadow(QFrame.Raised)
        v = QVBoxLayout(self); v.setContentsMargins(12,10,12,10); v.setSpacing(8)
        header = QHBoxLayout(); header.setContentsMargins(0,0,0,0); header.setSpacing(6)
        self.title_lbl = QLabel(title); self.title_lbl.setObjectName("CardTitle")
        header.addWidget(self.title_lbl)
        header.addStretch(1)
        # إزالة زر الطي نهائياً
        self.toggle_btn = None
        v.addLayout(header)
        self.container = QWidget()
        self.container_layout = QVBoxLayout(self.container)
        self.container_layout.setContentsMargins(0,0,0,0)
        self.container_layout.setSpacing(6)
        v.addWidget(self.container)
        self.v = self.container_layout

# ================ حوار الإعدادات =================
class SettingsDialog(QDialog):
    def __init__(self, parent=None, config: Dict[str, Any]=None):
        super().__init__(parent)
        self.setWindowTitle(tr("settings_title"))
        self.setWindowIcon(icon_for_action("settings"))
        self.cfg = dict(config or {})
        lay = QVBoxLayout(self)

        # عام: لغة وثيم
        grp_general = QGroupBox(tr("settings_title"))
        g = QGridLayout(grp_general)
        g.setHorizontalSpacing(8); g.setVerticalSpacing(6)
        self.lang_combo = QComboBox(); self.lang_combo.addItems(["العربية","English"])
        self.lang_combo.setCurrentIndex(0 if self.cfg.get("lang","ar")=="ar" else 1)
        # قائمة الثيمات الموسّعة
        theme_items_ar = ["غامق (Dark)","فاتح (Light)",
                          "Ocean Breeze","Sunset Orange","Midnight Purple",
                          "Steel Gray","Forest Green","Ruby Red"]
        theme_items_en = ["Dark","Light",
                          "Ocean Breeze","Sunset Orange","Midnight Purple",
                          "Steel Gray","Forest Green","Ruby Red"]
        if self.cfg.get("lang","ar")=="ar":
            items = theme_items_ar
        else:
            items = theme_items_en
        self.theme_combo = QComboBox(); self.theme_combo.addItems(items)
        theme_key = self.cfg.get("theme","dark")
        theme_index_map = {
            "dark":0, "light":1, "ocean":2, "sunset":3, "midnight":4,
            "steel":5, "forest":6, "ruby":7
        }
        self.theme_combo.setCurrentIndex(theme_index_map.get(theme_key,0))
        g.addWidget(QLabel(tr("config_lang")), 0,0); g.addWidget(self.lang_combo, 0,1)
        g.addWidget(QLabel(tr("config_theme")), 1,0); g.addWidget(self.theme_combo, 1,1)

        # فلاتر ثانوية
        grp_filters = QGroupBox(tr("inputs"))
        f = QGridLayout(grp_filters)
        f.setHorizontalSpacing(8); f.setVerticalSpacing(6)
        self.vtype_combo = QComboBox()
        self.vtype_combo.addItems([
            tr("config_vt_all"), tr("config_vt_string"), tr("config_vt_expand"),
            tr("config_vt_multi"), tr("config_vt_dword"), tr("config_vt_qword"),
            tr("config_vt_binary")
        ])
        vtype_map_rev = {
            "all":0, "string":1, "expandstring":2, "multistring":3,
            "dword":4, "qword":5, "binary":6
        }
        self.vtype_combo.setCurrentIndex(vtype_map_rev.get(self.cfg.get("value_type","all"), 0))
        self.use_age = QCheckBox(tr("config_use_age")); self.use_age.setChecked(bool(self.cfg.get("use_age", False)))
        self.days_spin = QSpinBox(); self.days_spin.setRange(0,3650); self.days_spin.setValue(int(self.cfg.get("days", 7)))

        # استبدال حقل الحسابات بقائمة منسدلة ذكية
        self.accounts_combo = QComboBox()
        self._fill_accounts_combo()
        # تعيين القيمة الحالية
        owner_filter = self.cfg.get("owner_filter", "all").lower()
        owner_index_map = {"all":0, "systems":1, "localsystem":2, "users":3}
        self.accounts_combo.setCurrentIndex(owner_index_map.get(owner_filter, 0))

        f.addWidget(QLabel(tr("config_value_type")), 0,0); f.addWidget(self.vtype_combo, 0,1)
        f.addWidget(self.use_age, 1,0); f.addWidget(QLabel(tr("config_age")), 1,1); f.addWidget(self.days_spin, 1,2)
        f.addWidget(QLabel(tr("config_accounts")), 2,0); f.addWidget(self.accounts_combo, 2,1,1,2)

        # نسخ احتياطي
        grp_backup = QGroupBox(tr("config_backup"))
        b = QHBoxLayout(grp_backup); b.setContentsMargins(8,6,8,6)
        self.btn_backup_create = QPushButton(tr("backup_create")); self.btn_backup_create.setIcon(icon_for_action("backup_up"))
        self.btn_backup_restore = QPushButton(tr("backup_restore")); self.btn_backup_restore.setIcon(icon_for_action("backup_down"))
        b.addWidget(self.btn_backup_create); b.addWidget(self.btn_backup_restore); b.addStretch(1)

        lay.addWidget(grp_general)
        lay.addWidget(grp_filters)
        lay.addWidget(grp_backup)

        self.buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        self.buttons.button(QDialogButtonBox.Save).setText(tr("saved"))
        self.buttons.button(QDialogButtonBox.Cancel).setText(tr("cancel"))
        lay.addWidget(self.buttons)

        self.buttons.accepted.connect(self.accept)
        self.buttons.rejected.connect(self.reject)
        self.btn_backup_create.clicked.connect(self._do_backup)
        self.btn_backup_restore.clicked.connect(self._do_restore)

    def _fill_accounts_combo(self):
        items = [
            tr("owner_all"),
            tr("owner_systems"),
            tr("owner_localsystem"),
            tr("owner_users"),
        ]
        self.accounts_combo.clear()
        for label in items:
            self.accounts_combo.addItem(label)
        self.accounts_combo.setToolTip(tr("config_accounts"))

    def values(self) -> Dict[str, Any]:
        vtype_map = {
            0: "all", 1: "string", 2: "expandstring", 3: "multistring",
            4: "dword", 5: "qword", 6: "binary"
        }
        theme_map = {
            0: "dark", 1: "light", 2: "ocean", 3: "sunset",
            4: "midnight", 5: "steel", 6: "forest", 7: "ruby"
        }
        owner_filter_map = {
            0: "all", 1: "systems", 2: "localsystem", 3: "users"
        }
        return {
            "lang": "ar" if self.lang_combo.currentIndex()==0 else "en",
            "theme": theme_map.get(self.theme_combo.currentIndex(), "dark"),
            "value_type": vtype_map.get(self.vtype_combo.currentIndex(), "all"),
            "use_age": self.use_age.isChecked(),
            "days": self.days_spin.value(),
            "owner_filter": owner_filter_map.get(self.accounts_combo.currentIndex(), "all"),
        }

    def _do_backup(self):
        try:
            fname, _ = QFileDialog.getSaveFileName(self, tr("backup_create"), str(Path.home()), "JSON (*.json)")
            if not fname: return
            root = {
                "config": self.values(),
                "lists": self.parent()._current_lists_snapshot_kw(),
                "rules": self.parent().rules_meta,
            }
            Path(fname).write_text(json.dumps(root, ensure_ascii=False, indent=2), encoding="utf-8")
            QMessageBox.information(self, tr("settings_title"), tr("saved"))
        except Exception as e:
            QMessageBox.warning(self, tr("settings_title"), str(e))

    def _do_restore(self):
        try:
            fname, _ = QFileDialog.getOpenFileName(self, tr("backup_restore"), str(Path.home()), "JSON (*.json)")
            if not fname: return
            data = json.loads(Path(fname).read_text(encoding="utf-8"))
            cfg = data.get("config", {})
            lists = data.get("lists", {})
            rules = data.get("rules", [])
            self.lang_combo.setCurrentIndex(0 if cfg.get("lang","ar")=="ar" else 1)
            # تحديث عناصر الثيم في حال تغيّرت اللغة
            self.theme_combo.clear()
            if cfg.get("lang","ar")=="ar":
                self.theme_combo.addItems(["غامق (Dark)","فاتح (Light)",
                                           "Ocean Breeze","Sunset Orange","Midnight Purple",
                                           "Steel Gray","Forest Green","Ruby Red"])
            else:
                self.theme_combo.addItems(["Dark","Light",
                                           "Ocean Breeze","Sunset Orange","Midnight Purple",
                                           "Steel Gray","Forest Green","Ruby Red"])
            theme_index = {"dark":0,"light":1,"ocean":2,"sunset":3,"midnight":4,"steel":5,"forest":6,"ruby":7}.get(cfg.get("theme","dark"),0)
            self.theme_combo.setCurrentIndex(theme_index)
            vt = cfg.get("value_type","all")
            self.vtype_combo.setCurrentIndex({"all":0,"string":1,"expandstring":2,"multistring":3,"dword":4,"qword":5,"binary":6}.get(vt,0))
            self.use_age.setChecked(bool(cfg.get("use_age", False)))
            self.days_spin.setValue(int(cfg.get("days",7)))

            # استرجاع عامل فلترة المالك
            self._fill_accounts_combo()
            owner_filter = cfg.get("owner_filter", "all").lower()
            owner_index_map = {"all":0, "systems":1, "localsystem":2, "users":3}
            self.accounts_combo.setCurrentIndex(owner_index_map.get(owner_filter, 0))

            self.parent()._restore_lists_and_rules(lists, rules)
            QMessageBox.information(self, tr("settings_title"), tr("loaded"))
        except Exception as e:
            QMessageBox.warning(self, tr("settings_title"), str(e))

# ================ مستعرض السجل (اختيار متعدد) ================
class RegistryBrowserDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(tr("browse_registry_title"))
        self.setWindowIcon(icon_for_action("browse"))
        self.resize(800, 600)
        v = QVBoxLayout(self)
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels([ "Name", "Path" ] if LANG=="en" else ["الاسم","المسار"])
        self.tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.tree.setColumnWidth(0, 320)
        v.addWidget(self.tree)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.button(QDialogButtonBox.Ok).setText(tr("ok"))
        btns.button(QDialogButtonBox.Cancel).setText(tr("cancel"))
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        v.addWidget(btns)

        self.tree.itemExpanded.connect(self._on_expand)
        self._populate_roots()
        self.selected_paths: List[str] = []

    def _populate_roots(self):
        for name in ["HKLM","HKCU","HKCR","HKU","HKCC"]:
            item = QTreeWidgetItem([name, name])
            item.setData(0, Qt.UserRole, name)
            item.addChild(QTreeWidgetItem(["(loading)...",""]))
            self.tree.addTopLevelItem(item)

    def _on_expand(self, item: QTreeWidgetItem):
        if item.childCount()==1 and item.child(0).text(0).startswith("("):
            item.removeChild(item.child(0))
            full = item.text(1)
            try:
                h, subkey = parse_registry_path(full)
                if h is None: return
                access = winreg.KEY_READ
                flags = [access]
                if hasattr(winreg,"KEY_WOW64_64KEY"):
                    flags += [access|winreg.KEY_WOW64_64KEY, access|winreg.KEY_WOW64_32KEY]
                opened = None
                for flg in flags:
                    try:
                        opened = winreg.OpenKey(h, subkey, 0, flg)
                        break
                    except Exception:
                        continue
                if opened is None:
                    item.addChild(QTreeWidgetItem(["[Access Denied]",""]))
                    return
                try:
                    count = winreg.QueryInfoKey(opened)[0]
                except Exception:
                    count = 0
                for i in range(count):
                    try:
                        name = winreg.EnumKey(opened, i)
                        child_path = full + ("\\" if subkey or full else "\\") + name
                        ch = QTreeWidgetItem([name, child_path])
                        ch.addChild(QTreeWidgetItem(["(loading)...",""]))
                        item.addChild(ch)
                    except Exception:
                        continue
                try:
                    winreg.CloseKey(opened)
                except Exception:
                    pass
            except PermissionError:
                item.addChild(QTreeWidgetItem(["[Access Denied]",""]))
            except Exception:
                item.addChild(QTreeWidgetItem(["[Error]",""]))

    def accept(self):
        self.selected_paths = []
        for it in self.tree.selectedItems():
            p = it.text(1)
            if p and not p.startswith("[") and not p.endswith("(loading)..."):
                self.selected_paths.append(p)
        super().accept()

# ================ حوار إدخال عام مع تلميح ================
class LabeledInputDialog(QDialog):
    def __init__(self, title: str, label: str, placeholder: str = "", text: str = "", parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setWindowIcon(icon_for_action("edit"))
        self.resize(560, 200)
        v = QVBoxLayout(self)
        v.addWidget(QLabel(label))
        self.edit = QLineEdit(text)
        if placeholder:
            self.edit.setPlaceholderText(placeholder)
        v.addWidget(self.edit)
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.button(QDialogButtonBox.Ok).setText(tr("ok"))
        btns.button(QDialogButtonBox.Cancel).setText(tr("cancel"))
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        v.addWidget(btns)

    def getText(self) -> Tuple[str, bool]:
        ok = self.exec_() == QDialog.Accepted
        return self.edit.text(), ok

# ================ حوار تفاصيل نتيجة =================
class ResultDetailsDialog(QDialog):
    def __init__(self, item: Dict[str, Any], parent=None):
        super().__init__(parent)
        self.setWindowTitle(tr("result_details"))
        self.setWindowIcon(icon_for_action("info"))
        self.resize(780, 560)
        v = QVBoxLayout(self)
        fields = [
            ("Key", item.get("key","")),
            ("Property", item.get("value_name","")),
            ("Value", item.get("value_str","")),
            ("Matched keyword", item.get("matched_kw","")),
            ("Value type", item.get("value_type","")),
            ("Last modified", item.get("last_mod","")),
            ("Owner", item.get("owner","")),
            ("State", item.get("state","")),
            ("Matched rule", item.get("matched_rule","")),
            ("Reasons", ", ".join(item.get("reasons",[]))),
        ]
        labels_map_ar = {
            "Key":"المفتاح", "Property":"الخاصية", "Value":"القيمة",
            "Matched keyword":"الكلمة المطابقة", "Value type":"نوع القيمة",
            "Last modified":"آخر تعديل", "Owner":"المالك", "State":"الحالة",
            "Matched rule":"القاعدة المطابقة", "Reasons":"الأسباب",
        }
        for k, vval in fields:
            row = QHBoxLayout()
            lbl = QLabel(labels_map_ar.get(k,k) if LANG=="ar" else k + "")
            lbl.setStyleSheet("font-weight:600;")
            val_lbl = QLabel(str(vval))
            val_lbl.setTextInteractionFlags(Qt.TextSelectableByMouse)
            row.addWidget(lbl); row.addWidget(val_lbl); row.addStretch(1)
            v.addLayout(row)
        v.addWidget(QLabel("" if LANG=="en" else "القيمة (كاملة):"))
        txt = QTextEdit()
        txt.setReadOnly(True)
        txt.setPlainText(item.get("value_str",""))
        v.addWidget(txt, 1)
        btns = QDialogButtonBox(QDialogButtonBox.Close)
        btns.button(QDialogButtonBox.Close).setText(tr("ok"))
        btns.rejected.connect(self.reject); btns.accepted.connect(self.accept)
        v.addWidget(btns)

# ================ عنصر فلترة متقدّم ثابت =================
class AdvancedFilterWidget(QWidget):
    applied = pyqtSignal(dict)  # {"column": int or None, "mode": "partial|exact|regex", "text": str}

    def __init__(self, headers: List[str], parent=None, compact_for_statusbar: bool = False):
        super().__init__(parent)
        h = QHBoxLayout(self)
        h.setContentsMargins(0, 0, 0 if compact_for_statusbar else 4, 0)
        h.setSpacing(6)
        self.col_combo = QComboBox()
        self.col_combo.addItem(tr("filter_column"))
        for i, name in enumerate(headers):
            self.col_combo.addItem(name, i)
        self.mode_combo = QComboBox()
        self.mode_combo.addItems([tr("match_partial"), tr("match_exact"), tr("match_regex")])
        self.text_edit = QLineEdit()
        self.text_edit.setPlaceholderText(tr("filter_text"))
        self.apply_btn = QPushButton(tr("filter_apply"))
        self.apply_btn.setIcon(icon_for_action("apply"))
        # ضبط أحجام مدمجة لشريط الحالة
        if compact_for_statusbar:
            self.setFixedHeight(28)
            self.col_combo.setFixedHeight(24)
            self.mode_combo.setFixedHeight(24)
            self.text_edit.setFixedHeight(24)
            self.apply_btn.setFixedHeight(24)
            self.apply_btn.setIconSize(QSize(16,16))
            self.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
        else:
            self.apply_btn.setFixedHeight(28)
            self.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)

        # ترتيب
        h.addWidget(QLabel(tr("filter_column")))
        h.addWidget(self.col_combo)
        h.addWidget(QLabel(tr("filter_mode")))
        h.addWidget(self.mode_combo)
        h.addWidget(self.text_edit, 1)
        h.addWidget(self.apply_btn)

        self.apply_btn.clicked.connect(self._emit)

    def _emit(self):
        col_idx = self.col_combo.currentData()
        try:
            col_idx = int(col_idx) if col_idx is not None else None
        except Exception:
            col_idx = None
        mode_map = {
            0: "partial",
            1: "exact",
            2: "regex"
        }
        mode = mode_map.get(self.mode_combo.currentIndex(), "partial")
        text = self.text_edit.text()
        self.applied.emit({"column": col_idx, "mode": mode, "text": text})

    # تحديث عناوين (للترجمة الديناميكية)
    def retitle(self, headers: List[str]):
        current_col = self.col_combo.currentData()
        self.col_combo.blockSignals(True)
        self.col_combo.clear()
        self.col_combo.addItem(tr("filter_column"))
        for i, name in enumerate(headers):
            self.col_combo.addItem(name, i)
        if current_col is not None:
            idx = self.col_combo.findData(current_col)
            if idx >= 0:
                self.col_combo.setCurrentIndex(idx)
        self.col_combo.blockSignals(False)

        self.mode_combo.blockSignals(True)
        self.mode_combo.clear()
        self.mode_combo.addItems([tr("match_partial"), tr("match_exact"), tr("match_regex")])
        self.mode_combo.blockSignals(False)
        self.text_edit.setPlaceholderText(tr("filter_text"))
        self.apply_btn.setText(tr("filter_apply"))

# ================ الواجهة الرئيسية (تبويبية) ================
class Main(QMainWindow):
    def __init__(self):
        super().__init__()
        self.saved_keys: List[str] = []
        self.saved_kws: List[str] = []
        self.rules_meta: List[Dict[str, Any]] = []
        default_owner_filter = "users"  # الافتراضي: المستخدم الحالي
        self.config = {
            "lang":"ar", "theme":"dark",
            "value_type":"all", "use_age":False, "days":7,
            "owner_filter": default_owner_filter
        }

        # تحميل تهيئة/قوائم/قواعد
        self._load_config()
        self._load_lists()
        self._load_rules_meta()

        # حالات التبويب والنتائج
        self.current_scan_tab = "kw"
        self.scanner: Optional[RegistryScannerThread] = None

        self.last_kw: List[Dict[str, Any]] = []
        self.last_rules_res: List[Dict[str, Any]] = []
        self.last_criteria_kw: Optional[Criteria] = None
        self.last_rules_specs_kw: List[RuleSpec] = []
        self.last_criteria_rules: Optional[Criteria] = None
        self.last_rules_specs_rules: List[RuleSpec] = []

        # مؤقت لتحديث الواجهة أثناء المسح لسلاسة أعلى
        self.ui_heartbeat = QTimer(self)
        self.ui_heartbeat.setInterval(150)
        self.ui_heartbeat.timeout.connect(lambda: QApplication.processEvents())

        # بناء الواجهة
        self._build_ui()
        self._apply_theme_choice()
        self._apply_language()

    # ---------- بناء الواجهة
    def _build_ui(self):
        self.setWindowTitle(tr("title"))
        self.setWindowIcon(QIcon(pixmap_from_base64(SAFE_LOGO_BASE64)))
        self.resize(1600, 900)

        # شريط الأدوات
        self.toolbar = QToolBar()
        # منع إخفاء شريط الأدوات: تعطيل قائمة السياق وقابلية الإخفاء
        self.toolbar.setContextMenuPolicy(Qt.PreventContextMenu)
        self.toolbar.toggleViewAction().setVisible(False)
        self.addToolBar(Qt.TopToolBarArea, self.toolbar)

        def act(icon_name, text_key, object_name=None):
            a = QAction(icon_for_action(icon_name), tr(text_key), self)
            a.setIconText(tr(text_key))
            a.setToolTip(tr(text_key))
            if object_name:
                a.setObjectName(object_name)
            return a

        self.act_scan = act("scan","scan","act_scan")
        self.act_stop = act("stop","stop","act_stop")
        self.act_refresh = act("refresh","refresh","act_refresh")
        self.act_clear = act("clear","clear","act_clear")
        self.act_export = act("export","export","act_export")
        self.act_settings = act("settings","settings","act_settings")
        self.act_exit = act("exit","exit","act_exit")
        self.toolbar.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.toolbar.addAction(self.act_scan)
        self.toolbar.addAction(self.act_stop)
        self.toolbar.addAction(self.act_refresh)
        self.toolbar.addSeparator()
        self.toolbar.addAction(self.act_clear)
        self.toolbar.addAction(self.act_export)
        self.toolbar.addSeparator()
        self.toolbar.addAction(self.act_settings)
        self.toolbar.addAction(self.act_exit)

        # تبويبات
        self.tabs = QTabWidget()
        self.tabs.setDocumentMode(True)
        self.tabs.setTabPosition(QTabWidget.North)
        self.tabs.setObjectName("MainTabs")
        self.setCentralWidget(self.tabs)

        # تبويب الكلمات
        self._build_tab_keywords()
        # تبويب القواعد
        self._build_tab_rules()

        self.tabs.currentChanged.connect(self._on_tab_changed)

        # الحالة: تقسيم إلى جهتين بشكل واضح
        self.status = QStatusBar(); self.setStatusBar(self.status)

        # يسار: عدادات + تقدم داخل حاوية مستقلة
        left_container = QWidget()
        hl = QHBoxLayout(left_container)
        hl.setContentsMargins(6,0,6,0)
        hl.setSpacing(8)
        self.lbl_status_total = QLabel("Total: 0" if LANG=="en" else "الإجمالي: 0")
        self.lbl_status_susp = QLabel("Suspicious: 0" if LANG=="en" else "المشبوه: 0")
        self.lbl_status_rate = QLabel("0%")
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.progress.setMaximumWidth(220)
        self.progress.setFixedHeight(16)
        hl.addWidget(self.lbl_status_total)
        hl.addWidget(self.lbl_status_susp)
        hl.addWidget(self.lbl_status_rate)
        hl.addWidget(self.progress)
        hl.addStretch(1)
        self.status.addWidget(left_container, 1)

        # يمين: عنصر فلترة متقدّم داخل حاوية مستقلة ودائمة
        headers_kw_for_filter = [h for h in tr("tbl_headers") if h not in ([ "Matched rule" ] if LANG=="en" else ["القاعدة المطابقة"])]
        headers_rules_for_filter = [h for h in tr("tbl_headers") if h not in ([ "Matched keyword" ] if LANG=="en" else ["الكلمة المطابقة"])]

        self.filter_status = AdvancedFilterWidget(headers=headers_kw_for_filter, compact_for_statusbar=True)
        self._current_filter_headers_kw = headers_kw_for_filter
        self._current_filter_headers_rules = headers_rules_for_filter
        self.filter_status.applied.connect(self._apply_filter_statusbar)

        right_container = QWidget()
        hr = QHBoxLayout(right_container)
        hr.setContentsMargins(0,0,6,0)
        hr.setSpacing(6)
        hr.addStretch(1)
        hr.addWidget(self.filter_status)
        self.status.addPermanentWidget(right_container, 0)

        # إشارات شريط الأدوات
        self.act_scan.triggered.connect(self._start_scan)
        self.act_stop.triggered.connect(self._stop_scan_confirm)
        self.act_refresh.triggered.connect(self._refresh_last_scan)
        self.act_clear.triggered.connect(self._clear)
        self.act_export.triggered.connect(self._export)
        self.act_settings.triggered.connect(self._open_settings)
        self.act_exit.triggered.connect(self._exit_confirm)

    def _build_tab_keywords(self):
        tab = QWidget()
        layout = QVBoxLayout(tab); layout.setContentsMargins(10,10,10,6); layout.setSpacing(6)
        splitter = QSplitter(Qt.Vertical)
        layout.addWidget(splitter, 1)

        # أعلى: معايير + إحصاءات
        top = QWidget(); tlay = QHBoxLayout(top); tlay.setContentsMargins(0,0,0,0); tlay.setSpacing(6)

        # بطاقة المعايير
        self.card_inputs_kw = Card(tr("inputs"), collapsible=False)
        iv = self.card_inputs_kw.v
        grid_inputs = QGridLayout()
        grid_inputs.setHorizontalSpacing(8)
        grid_inputs.setVerticalSpacing(6)

        # وضع العرض
        display_box = QGroupBox(tr("display_modes"))
        dgrid = QGridLayout(display_box); dgrid.setHorizontalSpacing(6); dgrid.setVerticalSpacing(4)
        self.radio_display_matched_kw = QRadioButton(tr("display_matched"))
        self.radio_display_all_kw = QRadioButton(tr("display_all"))
        self.radio_display_matched_kw.setChecked(True)
        dgrid.addWidget(self.radio_display_matched_kw, 0,0)
        dgrid.addWidget(self.radio_display_all_kw, 0,1)

        # كلمات الفحص
        kw_box = QGroupBox(tr("keywords"))
        kwgrid = QGridLayout(kw_box); kwgrid.setHorizontalSpacing(6); kwgrid.setVerticalSpacing(4)
        self.kws_list_kw = QListWidget()
        self.kws_list_kw.setMaximumHeight(150)
        for w in self.saved_kws: self.kws_list_kw.addItem(QListWidgetItem(w))
        self.btn_add_kw = QPushButton(tr("add")); self.btn_add_kw.setIcon(icon_for_action("add"))
        self.btn_edit_kw = QPushButton(tr("edit")); self.btn_edit_kw.setIcon(icon_for_action("edit"))
        self.btn_rem_kw = QPushButton(tr("remove")); self.btn_rem_kw.setIcon(icon_for_action("remove"))
        kwgrid.addWidget(self.kws_list_kw, 0,0,3,1)
        kwgrid.addWidget(self.btn_add_kw, 0,1)
        kwgrid.addWidget(self.btn_edit_kw, 1,1)
        kwgrid.addWidget(self.btn_rem_kw, 2,1)

        # مفاتيح السجل
        keys_box = QGroupBox(tr("keys_list"))
        kgrid = QGridLayout(keys_box); kgrid.setHorizontalSpacing(6); kgrid.setVerticalSpacing(4)
        self.keys_list_kw = QListWidget(); self.keys_list_kw.setMaximumHeight(150)
        for k in self.saved_keys: self.keys_list_kw.addItem(QListWidgetItem(k))
        self.btn_add_key_kw = QPushButton(tr("add")); self.btn_add_key_kw.setIcon(icon_for_action("add"))
        self.btn_rem_key_kw = QPushButton(tr("remove")); self.btn_rem_key_kw.setIcon(icon_for_action("remove"))
        self.btn_browse_kw = QPushButton(tr("browse")); self.btn_browse_kw.setIcon(icon_for_action("browse"))
        kgrid.addWidget(self.keys_list_kw, 0,0,3,1)
        kgrid.addWidget(self.btn_add_key_kw, 0,1)
        kgrid.addWidget(self.btn_rem_key_kw, 1,1)
        kgrid.addWidget(self.btn_browse_kw, 2,1)

        grid_inputs.addWidget(display_box, 0, 0, 1, 2)
        grid_inputs.addWidget(keys_box, 1, 0, 1, 2)
        grid_inputs.addWidget(kw_box, 2, 0, 1, 2)
        iv.addLayout(grid_inputs)

        # بطاقة الإحصاءات
        self.card_stats_kw = Card(tr("stats"), collapsible=False)
        sv = self.card_stats_kw.v
        self.lbl_total_kw = QLabel("0"); self.lbl_susp_kw = QLabel("0"); self.lbl_rate_kw = QLabel("0%")
        meta = QHBoxLayout()
        meta.setContentsMargins(0,0,0,0); meta.setSpacing(10)
        meta.addWidget(QLabel("Total/الإجمالي:")); meta.addWidget(self.lbl_total_kw); meta.addStretch(1)
        meta.addWidget(QLabel("Suspicious/المشبوه:")); meta.addWidget(self.lbl_susp_kw); meta.addStretch(1)
        meta.addWidget(QLabel("Rate/النسبة:")); meta.addWidget(self.lbl_rate_kw)
        sv.addLayout(meta)
        sv.addWidget(QLabel(tr("reasons_chart")))
        if pg:
            self.plot_kw = pg.PlotWidget(); self.plot_kw.setMinimumHeight(180)
            sv.addWidget(self.plot_kw)
        else:
            self.plot_kw = None; sv.addWidget(QLabel(tr("note_chart")))

        tlay.addWidget(self.card_inputs_kw, 1)
        tlay.addWidget(self.card_stats_kw, 1)

        splitter.addWidget(top)

        # أسفل: نتائج
        self.card_results_kw = Card(tr("results"))
        rv = self.card_results_kw.v

        # جدول
        kw_headers = [h for h in tr("tbl_headers") if h not in ([ "Matched rule" ] if LANG=="en" else ["القاعدة المطابقة"])]
        self.table_kw = QTableWidget(0, len(kw_headers))
        self.table_kw.setHorizontalHeaderLabels(kw_headers)
        self.table_kw.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table_kw.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_kw.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table_kw.verticalHeader().setDefaultSectionSize(26)
        self.table_kw.horizontalHeader().setMinimumHeight(26)
        self.table_kw.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.table_kw.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.table_kw.setContextMenuPolicy(Qt.CustomContextMenu)
        rv.addWidget(self.table_kw, 1)

        splitter.addWidget(self.card_results_kw)
        splitter.setSizes([380, 520])

        # إشارات تبويب الكلمات
        self.table_kw.customContextMenuRequested.connect(lambda pos: self._show_table_context_menu(self.table_kw, self.last_kw, pos, table_kind="kw"))
        self.table_kw.itemDoubleClicked.connect(lambda it: self._open_result_details_row(self.last_kw, it))

        self.btn_add_kw.clicked.connect(self._add_kw)
        self.btn_edit_kw.clicked.connect(self._edit_kw)
        self.btn_rem_kw.clicked.connect(self._rem_kw)

        self.btn_add_key_kw.clicked.connect(lambda: self._add_key(self.keys_list_kw))
        self.btn_rem_key_kw.clicked.connect(lambda: self._rem_key(self.keys_list_kw))
        self.btn_browse_kw.clicked.connect(lambda: self._browse_registry_into(self.keys_list_kw))

        self.keys_list_kw.installEventFilter(self)
        self.kws_list_kw.installEventFilter(self)

        # إضافة التبويب
        idx = self.tabs.addTab(tab, icon_for_action("tab_kw"), tr("tab_keywords"))
        self.tabs.setTabToolTip(idx, tr("tab_keywords"))

    def _build_tab_rules(self):
        tab = QWidget()
        layout = QVBoxLayout(tab); layout.setContentsMargins(10,10,10,6); layout.setSpacing(6)
        splitter = QSplitter(Qt.Vertical)
        layout.addWidget(splitter, 1)

        # أعلى: إدارة القواعد + مفاتيح القواعد + إحصاءات
        top = QWidget(); tlay = QHBoxLayout(top); tlay.setContentsMargins(0,0,0,0); tlay.setSpacing(6)

        # بطاقة القواعد
        self.card_rules = Card(tr("rules_title"), collapsible=False)
        rgrid_out = QGridLayout()
        rgrid_out.setHorizontalSpacing(8); rgrid_out.setVerticalSpacing(6)
        self.rules_list_rules = QListWidget(); self.rules_list_rules.setMaximumHeight(180)
        for r in self.rules_meta:
            title = r.get("title") or Path(r.get("path","")).stem
            enabled = r.get("enabled", True)
            item = QListWidgetItem(f"[{'✓' if enabled else ' '}] {title}")
            item.setData(Qt.UserRole, r)
            self.rules_list_rules.addItem(item)
        # أزرار إدارة القواعد
        self.btn_rule_import = QPushButton(tr("rule_import")); self.btn_rule_import.setIcon(icon_for_action("file"))
        self.btn_rule_import_folder = QPushButton(tr("rule_import_folder")); self.btn_rule_import_folder.setIcon(icon_for_action("folder"))
        self.btn_rule_remove = QPushButton(tr("rule_remove")); self.btn_rule_remove.setIcon(icon_for_action("remove"))
        self.btn_rule_remove_all = QPushButton(tr("remove_all")); self.btn_rule_remove_all.setIcon(icon_for_action("delete"))
        self.lbl_yaml_note = QLabel("" if HAVE_YAML else tr("need_yaml"))

        # خيار عرض النتائج
        display_box_rules = QGroupBox(tr("display_modes"))
        dgrid_r = QGridLayout(display_box_rules); dgrid_r.setHorizontalSpacing(6); dgrid_r.setVerticalSpacing(4)
        self.radio_display_matched_rules = QRadioButton(tr("display_matched"))
        self.radio_display_all_rules = QRadioButton(tr("display_all"))
        self.radio_display_matched_rules.setChecked(True)
        dgrid_r.addWidget(self.radio_display_matched_rules, 0,0)
        dgrid_r.addWidget(self.radio_display_all_rules, 0,1)

        # مفاتيح مخصصة لفحص القواعد
        keys_box = QGroupBox(tr("keys_list_rules"))
        kgrid = QGridLayout(keys_box); kgrid.setHorizontalSpacing(6); kgrid.setVerticalSpacing(4)
        self.keys_list_rules = QListWidget(); self.keys_list_rules.setMaximumHeight(150)
        for k in self.saved_keys: self.keys_list_rules.addItem(QListWidgetItem(k))
        self.btn_add_key_r = QPushButton(tr("add")); self.btn_add_key_r.setIcon(icon_for_action("add"))
        self.btn_rem_key_r = QPushButton(tr("remove")); self.btn_rem_key_r.setIcon(icon_for_action("remove"))
        self.btn_browse_r = QPushButton(tr("browse")); self.btn_browse_r.setIcon(icon_for_action("browse"))
        kgrid.addWidget(self.keys_list_rules, 0,0,3,1)
        kgrid.addWidget(self.btn_add_key_r, 0,1)
        kgrid.addWidget(self.btn_rem_key_r, 1,1)
        kgrid.addWidget(self.btn_browse_r, 2,1)

        rgrid_out.addWidget(self.rules_list_rules, 0,0,5,1)
        rgrid_out.addWidget(self.btn_rule_import, 0,1)
        rgrid_out.addWidget(self.btn_rule_import_folder, 1,1)
        rgrid_out.addWidget(self.btn_rule_remove, 2,1)
        rgrid_out.addWidget(self.btn_rule_remove_all, 3,1)
        rgrid_out.addWidget(self.lbl_yaml_note, 4,1)
        rgrid_out.addWidget(display_box_rules, 5,0,1,2)
        rgrid_out.addWidget(keys_box, 6,0,1,2)
        self.card_rules.v.addLayout(rgrid_out)

        # بطاقة الإحصاءات
        self.card_stats_rules = Card(tr("stats"), collapsible=False)
        sv = self.card_stats_rules.v
        self.lbl_total_rules = QLabel("0"); self.lbl_susp_rules = QLabel("0"); self.lbl_rate_rules = QLabel("0%")
        meta = QHBoxLayout()
        meta.setContentsMargins(0,0,0,0); meta.setSpacing(10)
        meta.addWidget(QLabel("Total/الإجمالي:")); meta.addWidget(self.lbl_total_rules); meta.addStretch(1)
        meta.addWidget(QLabel("Suspicious/المشبوه:")); meta.addWidget(self.lbl_susp_rules); meta.addStretch(1)
        meta.addWidget(QLabel("Rate/النسبة:")); meta.addWidget(self.lbl_rate_rules)
        sv.addLayout(meta)
        sv.addWidget(QLabel(tr("reasons_chart")))
        if pg:
            self.plot_rules = pg.PlotWidget(); self.plot_rules.setMinimumHeight(180)
            sv.addWidget(self.plot_rules)
        else:
            self.plot_rules = None; sv.addWidget(QLabel(tr("note_chart")))

        tlay.addWidget(self.card_rules, 1)
        tlay.addWidget(self.card_stats_rules, 1)

        splitter.addWidget(top)

        # أسفل: نتائج
        self.card_results_rules = Card(tr("results"))
        rv = self.card_results_rules.v

        # الجدول
        rules_headers = [h for h in tr("tbl_headers") if h not in ([ "Matched keyword" ] if LANG=="en" else ["الكلمة المطابقة"])]
        self.table_rules = QTableWidget(0, len(rules_headers))
        self.table_rules.setHorizontalHeaderLabels(rules_headers)
        self.table_rules.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table_rules.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_rules.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table_rules.verticalHeader().setDefaultSectionSize(26)
        self.table_rules.horizontalHeader().setMinimumHeight(26)
        self.table_rules.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.table_rules.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.table_rules.setContextMenuPolicy(Qt.CustomContextMenu)
        rv.addWidget(self.table_rules, 1)

        splitter.addWidget(self.card_results_rules)
        splitter.setSizes([380, 520])

        # إشارات تبويب القواعد
        self.table_rules.customContextMenuRequested.connect(lambda pos: self._show_table_context_menu(self.table_rules, self.last_rules_res, pos, table_kind="rules"))
        self.table_rules.itemDoubleClicked.connect(lambda it: self._open_result_details_row(self.last_rules_res, it))

        self.rules_list_rules.installEventFilter(self)
        self.keys_list_rules.installEventFilter(self)

        self.btn_rule_import.clicked.connect(self._import_rules_dialog)
        self.btn_rule_import_folder.clicked.connect(self._import_rules_folder_dialog)
        self.btn_rule_remove.clicked.connect(self._remove_selected_rules)
        self.btn_rule_remove_all.clicked.connect(self._remove_all_rules)
        self.rules_list_rules.itemDoubleClicked.connect(self._show_rule_details)

        self.btn_add_key_r.clicked.connect(lambda: self._add_key(self.keys_list_rules))
        self.btn_rem_key_r.clicked.connect(lambda: self._rem_key(self.keys_list_rules))
        self.btn_browse_r.clicked.connect(lambda: self._browse_registry_into(self.keys_list_rules))

        # إضافة التبويب
        idx = self.tabs.addTab(tab, icon_for_action("tab_rules"), tr("tab_rules"))
        self.tabs.setTabToolTip(idx, tr("tab_rules"))

    # ---------- أحداث عامة / تبويب
    def _on_tab_changed(self, index: int):
        self.current_scan_tab = "kw" if index == 0 else "rules"
        headers = (self._current_filter_headers_kw if self.current_scan_tab=="kw"
                   else self._current_filter_headers_rules)
        self.filter_status.retitle(headers)

    def closeEvent(self, event):
        try:
            self._save_lists()
            self._save_rules_meta()
        except Exception:
            pass
        super().closeEvent(event)

    def eventFilter(self, obj, event):
        if event.type() == QEvent.KeyPress and event.key() == Qt.Key_Delete:
            if obj is self.keys_list_kw:
                self._rem_key(self.keys_list_kw); return True
            elif obj is self.kws_list_kw:
                self._rem_kw(); return True
            elif obj is self.rules_list_rules:
                self._remove_selected_rules(); return True
            elif obj is self.keys_list_rules:
                self._rem_key(self.keys_list_rules); return True
        return super().eventFilter(obj, event)

    # ---------- إعدادات
    def _open_settings(self):
        dlg = SettingsDialog(self, config=self.config)
        if dlg.exec_() == QDialog.Accepted:
            self.config.update(dlg.values())
            self._save_config()
            self._apply_config()
            QMessageBox.information(self, tr("settings_title"), tr("config_saved"))

    def _apply_config(self):
        global LANG
        LANG = self.config.get("lang","ar")
        self._apply_language()
        self._apply_theme_choice()

    # ---------- لغة/ثيم
    def _apply_language(self):
        self.setWindowTitle(tr("title"))
        for act, key in [
            (self.act_scan,"scan"),(self.act_stop,"stop"),(self.act_refresh,"refresh"),
            (self.act_clear,"clear"),(self.act_export,"export"),
            (self.act_settings,"settings"),(self.act_exit,"exit")
        ]:
            act.setText(tr(key)); act.setIconText(tr(key)); act.setToolTip(tr(key))

        # تبويب الكلمات
        self.tabs.setTabText(0, tr("tab_keywords"))
        self.card_inputs_kw.title_lbl.setText(tr("inputs"))
        self.card_stats_kw.title_lbl.setText(tr("stats"))
        self.card_results_kw.title_lbl.setText(tr("results"))

        # أزرار تبويب الكلمات
        self.btn_add_kw.setText(tr("add"))
        self.btn_edit_kw.setText(tr("edit"))
        self.btn_rem_kw.setText(tr("remove"))
        self.btn_add_key_kw.setText(tr("add"))
        self.btn_rem_key_kw.setText(tr("remove"))
        self.btn_browse_kw.setText(tr("browse"))

        # عناصر عرض
        self.radio_display_matched_kw.setText(tr("display_matched"))
        self.radio_display_all_kw.setText(tr("display_all"))

        # تحديث عناوين الجدول (محذوف منه "Matched rule")
        kw_headers = [h for h in tr("tbl_headers") if h not in ([ "Matched rule" ] if LANG=="en" else ["القاعدة المطابقة"])]
        self.table_kw.setHorizontalHeaderLabels(kw_headers)
        self._current_filter_headers_kw = kw_headers

        # تبويب القواعد
        self.tabs.setTabText(1, tr("tab_rules"))
        self.card_rules.title_lbl.setText(tr("rules_title"))
        self.card_stats_rules.title_lbl.setText(tr("stats"))
        self.card_results_rules.title_lbl.setText(tr("results"))

        # أزرار القواعد
        self.btn_rule_import.setText(tr("rule_import"))
        self.btn_rule_import_folder.setText(tr("rule_import_folder"))
        self.btn_rule_remove.setText(tr("rule_remove"))
        self.btn_rule_remove_all.setText(tr("remove_all"))
        self.lbl_yaml_note.setText("" if HAVE_YAML else tr("need_yaml"))

        self.radio_display_matched_rules.setText(tr("display_matched"))
        self.radio_display_all_rules.setText(tr("display_all"))

        # تحديث عناوين الجدول (محذوف منه "Matched keyword")
        rules_headers = [h for h in tr("tbl_headers") if h not in ([ "Matched keyword" ] if LANG=="en" else ["الكلمة المطابقة"])]
        self.table_rules.setHorizontalHeaderLabels(rules_headers)
        self._current_filter_headers_rules = rules_headers

        # شريط الحالة (تحديث عدادات النصوص)
        self._update_status_counts(0,0,0.0)

        # تحديث شريط الفلترة وفق الجدول النشط
        headers_for_active = self._current_filter_headers_kw if self.current_scan_tab=="kw" else self._current_filter_headers_rules
        self.filter_status.retitle(headers_for_active)

    def _apply_theme_choice(self):
        theme = self.config.get("theme","dark")
        if theme == "light":
            self._apply_light()
        elif theme == "ocean":
            self._apply_theme_custom(bg="#0b1c2c", panel="#11263a", card1="#15344d", text="#e6f4ff",
                                     accent="#35a8ff", header="#1d3a56", tab_active="#1e4a6b", tab_inactive="#15344d",
                                     btn_scan="#22c55e", btn_stop="#ef4444", btn_export="#8b5cf6")
        elif theme == "sunset":
            self._apply_theme_custom(bg="#28150f", panel="#331a12", card1="#3d2016", text="#ffeedd",
                                     accent="#ff8a3d", header="#4a261a", tab_active="#5a2e1f", tab_inactive="#3d2016",
                                     btn_scan="#fb923c", btn_stop="#f87171", btn_export="#f59e0b")
        elif theme == "midnight":
            self._apply_theme_custom(bg="#120b1a", panel="#190e24", card1="#1f1030", text="#efe6ff",
                                     accent="#8b5cf6", header="#2a1842", tab_active="#321f4d", tab_inactive="#1f1030",
                                     btn_scan="#22c55e", btn_stop="#ef4444", btn_export="#8b5cf6")
        elif theme == "steel":
            self._apply_theme_custom(bg="#141618", panel="#1b1e22", card1="#21252b", text="#e8edf2",
                                     accent="#7f8c98", header="#262b32", tab_active="#2a3038", tab_inactive="#21252b",
                                     btn_scan="#22c55e", btn_stop="#ef4444", btn_export="#8b5cf6")
        elif theme == "forest":
            self._apply_theme_custom(bg="#0e1a12", panel="#102016", card1="#14281c", text="#e6ffe9",
                                     accent="#34d399", header="#173524", tab_active="#1b452c", tab_inactive="#14281c",
                                     btn_scan="#22c55e", btn_stop="#ef4444", btn_export="#10b981")
        elif theme == "ruby":
            self._apply_theme_custom(bg="#20080c", panel="#2a0b10", card1="#330d14", text="#ffe6ea",
                                     accent="#ef4444", header="#3d0f18", tab_active="#4a1120", tab_inactive="#330d14",
                                     btn_scan="#f97316", btn_stop="#ef4444", btn_export="#e11d48")
        else:
            self._apply_dark()

    def _apply_dark(self):
        app = QApplication.instance()
        app.setStyleSheet("""
        QWidget { font-family: "Segoe UI", Tahoma; font-size: 10pt; color: #e6e6e6; }
        QMainWindow, QDialog, QMenu { background: #0f1220; }
        #Card {
            background: qlineargradient(x1:0,y1:0,x2:1,y2:1, stop:0 #1c2236, stop:1 #121729);
            border: 1px solid #2f3b5a; border-radius: 12px;
        }
        #CardTitle { font-weight:600; font-size: 11pt; color:#a8c1ff; }
        QLabel { color:#e6e6e6; }
        QLineEdit, QComboBox, QListWidget, QSpinBox, QTreeWidget, QTextEdit, QRadioButton {
            background:#0c1020; border:1px solid #334062; border-radius:6px; padding:4px; color:#e6e6e6;
            min-height: 22px;
        }
        QGroupBox { border: 1px solid #2f3b5a; border-radius:8px; margin-top:6px; padding:6px; }
        QGroupBox::title { subcontrol-origin: margin; left: 8px; padding: 0 4px; color:#9bb3ff; }
        QTableWidget, QTreeWidget {
            background:#0c1020; gridline-color:#2d395a; border:1px solid #2f3b5a; border-radius:8px;
        }
        QHeaderView::section {
            background:#172038; color:#cfe3ff; border:0px; padding:6px; border-right:1px solid #2f3b5a;
            font-weight:600; min-height:24px;
        }
        QTableWidget::item:selected { background:#23345d; }
        QToolBar { background:#121729; border:0px; padding:4px; }
        QStatusBar { background:#121729; border-top:1px solid #2f3b5a; }
        QProgressBar { background:#0c1020; border:1px solid #334062; border-radius:6px; text-align:center; color:#cfe3ff; }
        QProgressBar::chunk { background:#3a86ff; border-radius:6px; }
        QPushButton {
            background:#1b2240; border:2px solid #334062; border-radius:6px; padding:4px 10px; color:#e6e6e6;
        }
        QPushButton:hover { background:#233054; }
        QMessageBox { background:#0f1220; }

        QTabBar::tab {
            background:#1b2240; color:#cfe3ff; padding:6px 12px; border-top-left-radius:8px; border-top-right-radius:8px; margin-right:4px;
            border:1px solid #2f3b5a;
        }
        QTabBar::tab:selected {
            background:#2a3560; color:#ffffff; font-weight:700; border:2px solid #4b5b8a;
        }
        QTabBar::tab:!selected { font-weight:500; }
        """)

        # إبراز الأزرار الرئيسية
        self._style_primary_buttons(scan="#22c55e", stop="#ef4444", export="#8b5cf6")

        # ألوان الرسوم
        if getattr(self, 'plot_kw', None):
            self.plot_kw.setBackground('#0f1220')
            for axis in ('left','bottom'):
                self.plot_kw.getAxis(axis).setPen('#9bb3ff')
                self.plot_kw.getAxis(axis).setTextPen('#cfe3ff')
        if getattr(self, 'plot_rules', None):
            self.plot_rules.setBackground('#0f1220')
            for axis in ('left','bottom'):
                self.plot_rules.getAxis(axis).setPen('#9bb3ff')
                self.plot_rules.getAxis(axis).setTextPen('#cfe3ff')

    def _apply_light(self):
        app = QApplication.instance()
        app.setStyleSheet("""
        QWidget { font-family: "Segoe UI", Tahoma; font-size: 10pt; color: #1c2030; }
        QMainWindow, QDialog, QMenu { background: #f5f7fb; }
        #Card { background: #ffffff; border: 1px solid #e3e8f5; border-radius: 12px; }
        #CardTitle { font-weight:600; font-size: 11pt; color:#3b5bcc; }
        QLineEdit, QComboBox, QListWidget, QSpinBox, QTreeWidget, QTextEdit, QRadioButton {
            background:#ffffff; border:1px solid #cfd7ee; border-radius:6px; padding:4px; color:#222;
            min-height: 22px;
        }
        QGroupBox { border: 1px solid #e3e8f5; border-radius:8px; margin-top:6px; padding:6px; }
        QGroupBox::title { subcontrol-origin: margin; left: 8px; padding: 0 4px; color:#3b5bcc; }
        QTableWidget, QTreeWidget { background:#ffffff; gridline-color:#dfe6f7; border:1px solid #e3e8f5; border-radius:8px; }
        QHeaderView::section { background:#eef2ff; color:#2d3c77; border:0px; padding:6px; border-right:1px solid #dde5fb; font-weight:600; min-height:24px;}
        QTableWidget::item:selected { background:#dfe6ff; }
        QToolBar { background:#eef2ff; border:0px; padding:4px; }
        QStatusBar { background:#eef2ff; border-top:1px solid #dde5fb; }
        QProgressBar { background:#ffffff; border:1px solid #cfd7ee; border-radius:6px; text-align:center; color:#1c2030; }
        QProgressBar::chunk { background:#4f6fff; border-radius:6px; }
        QPushButton { background:#ffffff; border:2px solid #cfd7ee; border-radius:6px; padding:4px 10px; color:#1c2030; }
        QPushButton:hover { background:#eef2ff; }
        QMessageBox { background:#f5f7fb; }

        QTabBar::tab {
            background:#ffffff; color:#2d3c77; padding:6px 12px; border-top-left-radius:8px; border-top-right-radius:8px; margin-right:4px;
            border:1px solid #e3e8f5;
        }
        QTabBar::tab:selected {
            background:#dfe6ff; color:#1c2030; font-weight:700; border:2px solid #b9c6ff;
        }
        QTabBar::tab:!selected { font-weight:500; }
        """)
        self._style_primary_buttons(scan="#22c55e", stop="#ef4444", export="#3b82f6")
        if getattr(self, 'plot_kw', None):
            self.plot_kw.setBackground('w')
            for axis in ('left','bottom'):
                self.plot_kw.getAxis(axis).setPen('#6b7bb7')
                self.plot_kw.getAxis(axis).setTextPen('#2d3c77')
        if getattr(self, 'plot_rules', None):
            self.plot_rules.setBackground('w')
            for axis in ('left','bottom'):
                self.plot_rules.getAxis(axis).setPen('#6b7bb7')
                self.plot_rules.getAxis(axis).setTextPen('#2d3c77')

    def _apply_theme_custom(self, bg, panel, card1, text, accent, header,
                             tab_active, tab_inactive, btn_scan, btn_stop, btn_export):
        app = QApplication.instance()
        style = f"""
        QWidget {{ font-family: "Segoe UI", Tahoma; font-size: 10pt; color: {text}; }}
        QMainWindow, QDialog, QMenu {{ background: {bg}; }}
        #Card {{
            background: {card1};
            border: 1px solid {accent}; border-radius: 12px;
        }}
        #CardTitle {{ font-weight:700; font-size: 11pt; color:{text}; }}
        QLabel {{ color:{text}; }}
        QLineEdit, QComboBox, QListWidget, QSpinBox, QTreeWidget, QTextEdit, QRadioButton {{
            background:{panel}; border:1px solid {accent}; border-radius:6px; padding:4px; color:{text};
            min-height: 22px;
        }}
        QGroupBox {{ border: 1px solid {accent}; border-radius:8px; margin-top:6px; padding:6px; }}
        QGroupBox::title {{ subcontrol-origin: margin; left: 8px; padding: 0 4px; color:{text}; }}
        QTableWidget, QTreeWidget {{
            background:{panel}; gridline-color:{accent}; border:1px solid {accent}; border-radius:8px;
        }}
        QHeaderView::section {{
            background:{header}; color:{text}; border:0px; padding:6px; border-right:1px solid {accent}; font-weight:700; min-height:24px;
        }}
        QTableWidget::item:selected {{ background:{tab_active}; }}
        QToolBar {{ background:{header}; border:0px; padding:4px; }}
        QStatusBar {{ background:{header}; border-top:1px solid {accent}; }}
        QProgressBar {{ background:{panel}; border:1px solid {accent}; border-radius:6px; text-align:center; color:{text}; }}
        QProgressBar::chunk {{ background:{accent}; border-radius:6px; }}
        QPushButton {{
            background:{panel}; border:2px solid {accent}; border-radius:6px; padding:4px 10px; color:{text};
        }}
        QPushButton:hover {{ background:{tab_inactive}; }}
        QMessageBox {{ background:{bg}; }}

        QTabBar::tab {{
            background:{tab_inactive}; color:{text}; padding:6px 12px; border-top-left-radius:8px; border-top-right-radius:8px; margin-right:4px;
            border:1px solid {accent};
        }}
        QTabBar::tab:selected {{
            background:{tab_active}; color:{text}; font-weight:700; border:2px solid {accent};
        }}
        QTabBar::tab:!selected {{ font-weight:500; }}
        """
        app.setStyleSheet(style)
        self._style_primary_buttons(scan=btn_scan, stop=btn_stop, export=btn_export)

        if getattr(self, 'plot_kw', None):
            self.plot_kw.setBackground(bg)
            for axis in ('left','bottom'):
                self.plot_kw.getAxis(axis).setPen(text)
                self.plot_kw.getAxis(axis).setTextPen(text)
        if getattr(self, 'plot_rules', None):
            self.plot_rules.setBackground(bg)
            for axis in ('left','bottom'):
                self.plot_rules.getAxis(axis).setPen(text)
                self.plot_rules.getAxis(axis).setTextPen(text)

    def _style_primary_buttons(self, scan="#22c55e", stop="#ef4444", export="#8b5cf6"):
        self.toolbar.setStyleSheet(f"""
        QToolBar {{ padding:4px; }}
        QToolButton#act_scan {{ background:{scan}22; border:2px solid {scan}; border-radius:6px; padding:4px 8px; }}
        QToolButton#act_stop {{ background:{stop}22; border:2px solid {stop}; border-radius:6px; padding:4px 8px; }}
        QToolButton#act_export {{ background:{export}22; border:2px solid {export}; border-radius:6px; padding:4px 8px; }}
        QToolButton:hover {{ opacity:0.95; }}
        """)

    # ---------- إدارة القوائم المشتركة
    def _add_key(self, list_widget: QListWidget):
        dlg = LabeledInputDialog(tr("add_key_title"), tr("add_key_hint"), tr("add_key_hint"), "", self)
        text, ok = dlg.getText()
        if ok and text.strip():
            vals = [list_widget.item(i).text() for i in range(list_widget.count())]
            if text.strip() not in vals: list_widget.addItem(QListWidgetItem(text.strip()))

    def _rem_key(self, list_widget: QListWidget):
        sels = list_widget.selectedItems()
        if not sels:
            return
        if self._confirm(tr("confirm_delete_selected")):
            for it in sels:
                list_widget.takeItem(list_widget.row(it))

    def _browse_registry_into(self, list_widget: QListWidget):
        dlg = RegistryBrowserDialog(self)
        if dlg.exec_() == QDialog.Accepted and dlg.selected_paths:
            vals = [list_widget.item(i).text() for i in range(list_widget.count())]
            for path in dlg.selected_paths:
                if path not in vals:
                    list_widget.addItem(QListWidgetItem(path))

    def _add_kw(self):
        dlg = LabeledInputDialog(tr("add_kw_title"), tr("add_kw_hint"), tr("add_kw_hint"), "", self)
        text, ok = dlg.getText()
        if ok and text.strip():
            items = [t.strip() for t in text.split(",") if t.strip()]
            vals = [self.kws_list_kw.item(i).text() for i in range(self.kws_list_kw.count())]
            for t in items:
                if t not in vals:
                    self.kws_list_kw.addItem(QListWidgetItem(t))

    def _edit_kw(self):
        it = self.kws_list_kw.currentItem()
        if not it:
            return
        dlg = LabeledInputDialog(tr("edit_kw_title"), tr("add_kw_hint"), "", it.text(), self)
        text, ok = dlg.getText()
        if ok and text.strip():
            it.setText(text.strip())

    def _rem_kw(self):
        sels = self.kws_list_kw.selectedItems()
        if not sels: return
        if self._confirm(tr("confirm_delete_selected")):
            for it in sels:
                self.kws_list_kw.takeItem(self.kws_list_kw.row(it))

    # حفظ تلقائي للقوائم (مشتركة بين التبويبين)
    def _save_lists(self):
        data = {
            "keys": [self.keys_list_kw.item(i).text() for i in range(self.keys_list_kw.count())],
            "kws": [self.kws_list_kw.item(i).text() for i in range(self.kws_list_kw.count())]
        }
        try:
            with open(LISTS_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _load_lists(self):
        if LISTS_FILE.exists():
            try:
                data = json.load(open(LISTS_FILE, "r", encoding="utf-8"))
                self.saved_keys = data.get("keys", ["HKLM\\SOFTWARE"])
                self.saved_kws = data.get("kws", ["ssh","reverse"])
            except Exception:
                self.saved_keys = ["HKLM\\SOFTWARE"]; self.saved_kws = ["ssh","reverse"]
        else:
            self.saved_keys = ["HKLM\\SOFTWARE"]; self.saved_kws = ["ssh","reverse"]

    def _current_lists_snapshot_kw(self) -> Dict[str, List[str]]:
        return {
            "keys": [self.keys_list_kw.item(i).text() for i in range(self.keys_list_kw.count())],
            "kws": [self.kws_list_kw.item(i).text() for i in range(self.kws_list_kw.count())]
        }

    def _restore_lists_and_rules(self, lists: Dict[str, Any], rules: List[Dict[str, Any]]):
        try:
            # قوائم تبويب الكلمات
            self.keys_list_kw.clear(); self.kws_list_kw.clear()
            for k in lists.get("keys", []): self.keys_list_kw.addItem(QListWidgetItem(k))
            for w in lists.get("kws", []): self.kws_list_kw.addItem(QListWidgetItem(w))
            # قوائم مفاتيح تبويب القواعد = نفس المفاتيح لسهولة
            self.keys_list_rules.clear()
            for k in lists.get("keys", []): self.keys_list_rules.addItem(QListWidgetItem(k))
            # القواعد
            self.rules_meta = []
            for r in rules:
                self.rules_meta.append({
                    "path": r.get("path",""),
                    "enabled": bool(r.get("enabled", True)),
                    "title": r.get("title") or Path(r.get("path","")).stem,
                    "level": r.get("level",""),
                })
            self._rebuild_rules_list_widget()
            self._save_rules_meta()
            self._save_lists()
        except Exception:
            pass

    # ---------- القواعد
    def _load_rules_meta(self):
        if RULES_FILE.exists():
            try:
                data = json.load(open(RULES_FILE, "r", encoding="utf-8"))
                self.rules_meta = []
                for r in data:
                    self.rules_meta.append({
                        "path": r.get("path",""),
                        "enabled": bool(r.get("enabled", True)),
                        "title": r.get("title") or Path(r.get("path","")).stem,
                        "level": r.get("level",""),
                    })
            except Exception:
                self.rules_meta = []
        else:
            self.rules_meta = []

    def _save_rules_meta(self):
        try:
            with open(RULES_FILE, "w", encoding="utf-8") as f:
                json.dump(self.rules_meta, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _rebuild_rules_list_widget(self):
        self.rules_list_rules.clear()
        for r in self.rules_meta:
            title = r.get("title") or Path(r.get("path","")).stem
            enabled = r.get("enabled", True)
            item = QListWidgetItem(f"[{'✓' if enabled else ' '}] {title}")
            item.setData(Qt.UserRole, r)
            self.rules_list_rules.addItem(item)

    def _collect_rules_for_scanning_from_rules_tab(self) -> List[Dict[str, Any]]:
        rules = []
        for i in range(self.rules_list_rules.count()):
            meta = self.rules_list_rules.item(i).data(Qt.UserRole) or {}
            rules.append(meta)
        return rules

    def _import_rules_dialog(self):
        files, _ = QFileDialog.getOpenFileNames(self, tr("rule_import"), str(Path.home()), "YAML (*.yml *.yaml)")
        if not files:
            return
        self._import_rule_files(files)

    def _import_rules_folder_dialog(self):
        folder = QFileDialog.getExistingDirectory(self, tr("rule_import_folder"), str(Path.home()))
        if not folder:
            return
        p = Path(folder)
        files = [str(fp) for fp in p.rglob("*") if fp.suffix.lower() in (".yml",".yaml")]
        if not files:
            return
        self._import_rule_files(files)

    def _import_rule_files(self, files: List[str]):
        # مؤشر تقدّم أثناء الاستيراد
        progress = QProgressDialog(tr("progress"), tr("cancel"), 0, len(files), self)
        progress.setWindowTitle(tr("rule_import"))
        progress.setWindowModality(Qt.ApplicationModal)
        progress.setMinimumDuration(0)
        progress.setAutoClose(True)
        progress.setAutoReset(True)

        added = 0
        for idx, fp in enumerate(files, start=1):
            if progress.wasCanceled():
                break
            progress.setLabelText(f"{tr('rule_import')}: {Path(fp).name}")
            progress.setValue(idx-1)
            QApplication.processEvents()
            try:
                p = Path(fp)
                if not p.exists():
                    continue
                title, level = "", ""
                if HAVE_YAML:
                    try:
                        data = yaml.safe_load(p.read_text(encoding="utf-8", errors="ignore")) or {}
                        title = data.get("title") or data.get("id") or p.stem
                        level = data.get("level","")
                    except Exception:
                        title = p.stem
                        level = ""
                else:
                    title = p.stem
                    level = ""
                meta = {"path": str(p), "enabled": True, "title": title, "level": level}
                exists = any(r.get("path","").lower()==str(p).lower() for r in self.rules_meta)
                if not exists:
                    self.rules_meta.append(meta)
                    added += 1
            except Exception:
                continue

        progress.setValue(len(files))
        if added:
            self._save_rules_meta()
            self._rebuild_rules_list_widget()
        QMessageBox.information(self, tr("rules_title"), f"{tr('loaded')}: {added}")

    def _remove_selected_rules(self):
        sels = self.rules_list_rules.selectedItems()
        if not sels: return
        if not self._confirm(tr("confirm_delete_selected")):
            return
        paths_to_remove = set()
        for it in sels:
            meta = it.data(Qt.UserRole) or {}
            p = meta.get("path","")
            if p: paths_to_remove.add(p)
        self.rules_meta = [r for r in self.rules_meta if r.get("path","") not in paths_to_remove]
        self._save_rules_meta()
        self._rebuild_rules_list_widget()

    def _remove_all_rules(self):
        if not self.rules_meta:
            return
        if not self._confirm(tr("confirm_delete_all_rules")):
            return
        self.rules_meta = []
        self._save_rules_meta()
        self._rebuild_rules_list_widget()

    def _show_rule_details(self, item: QListWidgetItem):
        meta = item.data(Qt.UserRole) or {}
        path = meta.get("path","")
        title = meta.get("title","")
        level = meta.get("level","")
        text = f"{tr('rule_name')}: {title}\n{tr('rule_level')}: {level}\n{tr('rule_path')}: {path}\n\n"
        try:
            content = Path(path).read_text(encoding="utf-8", errors="ignore")
        except Exception:
            content = "[Failed to read rule file]"
        dlg = QDialog(self); dlg.setWindowTitle(tr("rule_details")); dlg.setWindowIcon(icon_for_action("file")); dlg.resize(760, 540)
        v = QVBoxLayout(dlg)
        cap = QLabel(text)
        v.addWidget(cap)
        editor = QTextEdit(); editor.setPlainText(content); editor.setReadOnly(True)
        v.addWidget(editor, 1)
        btns = QDialogButtonBox(QDialogButtonBox.Close)
        btns.button(QDialogButtonBox.Close).setText(tr("ok"))
        btns.rejected.connect(dlg.reject); btns.accepted.connect(dlg.accept)
        v.addWidget(btns)
        dlg.exec_()

    # ---------- تجميع المعايير
    def _criteria_keywords(self) -> Criteria:
        vtype = self.config.get("value_type","all")
        display_mode = "all" if self.radio_display_all_kw.isChecked() else "matched"
        return Criteria(
            keys=[self.keys_list_kw.item(i).text() for i in range(self.keys_list_kw.count())],
            keywords=[self.kws_list_kw.item(i).text() for i in range(self.kws_list_kw.count())],
            value_type=vtype,
            use_age=bool(self.config.get("use_age", False)),
            days=int(self.config.get("days", 0)),
            owner_filter=self.config.get("owner_filter", "all"),
            mode_keywords=True,
            mode_rules=False,
            display_mode=display_mode
        )

    def _criteria_rules(self) -> Criteria:
        vtype = self.config.get("value_type","all")
        display_mode = "all" if self.radio_display_all_rules.isChecked() else "matched"
        return Criteria(
            keys=[self.keys_list_rules.item(i).text() for i in range(self.keys_list_rules.count())],
            keywords=[],
            value_type=vtype,
            use_age=bool(self.config.get("use_age", False)),
            days=int(self.config.get("days", 0)),
            owner_filter=self.config.get("owner_filter", "all"),
            mode_keywords=False,
            mode_rules=True,
            display_mode=display_mode
        )

    # ---------- الفحص (تبويبي)
    def _start_scan(self):
        active_index = self.tabs.currentIndex()
        if active_index == 0:
            self.current_scan_tab = "kw"
            self._start_scan_keywords()
        else:
            self.current_scan_tab = "rules"
            self._start_scan_rules()
        headers = (self._current_filter_headers_kw if self.current_scan_tab=="kw"
                   else self._current_filter_headers_rules)
        self.filter_status.retitle(headers)

    def _start_scan_keywords(self):
        crit = self._criteria_keywords()
        if not crit.keys or not crit.keywords:
            QMessageBox.warning(self, tr("title"), tr("no_filters")); return
        self._begin_scan(crit, rules_specs=[])

    def _start_scan_rules(self):
        if not HAVE_YAML:
            QMessageBox.warning(self, tr("title"), tr("need_yaml")); return
        crit = self._criteria_rules()
        rules_meta = self._collect_rules_for_scanning_from_rules_tab()
        if not any(r.get("enabled", True) for r in rules_meta):
            QMessageBox.warning(self, tr("title"), tr("no_rules")); return
        rules_specs = load_rules_from_filelist(rules_meta)
        self._begin_scan(crit, rules_specs=rules_specs)

    def _begin_scan(self, crit: Criteria, rules_specs: List[RuleSpec]):
        # تهيئة واجهة التبويب النشط
        if self.current_scan_tab == "kw":
            self.table_kw.setRowCount(0)
        else:
            self.table_rules.setRowCount(0)
        self.progress.setVisible(True); self.progress.setRange(0,0)
        self.status.showMessage(tr("progress"))
        self.act_scan.setEnabled(False); self.act_stop.setEnabled(True); self.act_refresh.setEnabled(False)

        # تشغيل الماسح
        self.scanner = RegistryScannerThread(crit, rules=rules_specs)
        self.scanner.progress.connect(self._on_progress)
        self.scanner.finished.connect(self._on_finished_tabaware)
        self.scanner.error.connect(self._on_error)
        self.scanner.start()

        # تشغيل نبض تحديث الواجهة للتجاوب
        self.ui_heartbeat.start()

        if self.current_scan_tab == "kw":
            self.last_criteria_kw = crit
            self.last_rules_specs_kw = rules_specs
        else:
            self.last_criteria_rules = crit
            self.last_rules_specs_rules = rules_specs

    def _refresh_last_scan(self):
        if self.scanner and self.scanner.isRunning():
            return
        self.progress.setVisible(True); self.progress.setRange(0,0)
        self.status.showMessage(tr("progress"))
        self.act_scan.setEnabled(False); self.act_stop.setEnabled(True); self.act_refresh.setEnabled(False)
        if self.current_scan_tab == "kw":
            if not self.last_criteria_kw:
                QMessageBox.information(self, tr("title"), tr("no_filters")); self._stop_scan(); return
            self.table_kw.setRowCount(0)
            self.scanner = RegistryScannerThread(self.last_criteria_kw, rules=self.last_rules_specs_kw)
        else:
            if not self.last_criteria_rules:
                QMessageBox.information(self, tr("title"), tr("no_filters")); self._stop_scan(); return
            self.table_rules.setRowCount(0)
            self.scanner = RegistryScannerThread(self.last_criteria_rules, rules=self.last_rules_specs_rules)
        self.scanner.progress.connect(self._on_progress)
        self.scanner.finished.connect(self._on_finished_tabaware)
        self.scanner.error.connect(self._on_error)
        self.scanner.start()
        self.ui_heartbeat.start()

    def _stop_scan_confirm(self):
        if not self.scanner or not self.scanner.isRunning():
            return
        if not self._confirm(tr("confirm_stop_scan")):
            return
        self._stop_scan()

    def _stop_scan(self):
        if self.scanner:
            self.scanner.stop()
        self.act_scan.setEnabled(True); self.act_stop.setEnabled(False); self.act_refresh.setEnabled(True)
        self.progress.setVisible(False); self.status.clearMessage()
        self.ui_heartbeat.stop()

    def _clear(self):
        if self.current_scan_tab == "kw":
            self.table_kw.setRowCount(0); self.last_kw = []
            self.lbl_total_kw.setText("0"); self.lbl_susp_kw.setText("0"); self.lbl_rate_kw.setText("0%")
            if getattr(self, 'plot_kw', None): self.plot_kw.clear()
            self._update_status_counts(0,0,0.0)
        else:
            self.table_rules.setRowCount(0); self.last_rules_res = []
            self.lbl_total_rules.setText("0"); self.lbl_susp_rules.setText("0"); self.lbl_rate_rules.setText("0%")
            if getattr(self, 'plot_rules', None): self.plot_rules.clear()
            self._update_status_counts(0,0,0.0)
        self.status.clearMessage()

    def _exit_confirm(self):
        if not self._confirm(tr("confirm_exit")):
            return
        QApplication.instance().quit()

    def _on_progress(self, count:int):
        self.status.showMessage(f"{tr('progress')} ({count})")

    def _on_finished_tabaware(self, items: List[Dict[str,Any]], total:int):
        try:
            if self.current_scan_tab == "kw":
                self.last_kw = items
                self._fill_table_and_stats_kw(items, total)
            else:
                self.last_rules_res = items
                self._fill_table_and_stats_rules(items, total)
            self.act_scan.setEnabled(True); self.act_stop.setEnabled(False); self.act_refresh.setEnabled(True)
            self.progress.setVisible(False); self.progress.setRange(0,100)
            self.ui_heartbeat.stop()
        except Exception as e:
            traceback.print_exc()
            self._on_error(str(e))

    def _fill_table_and_stats_kw(self, items, total):
        matched_count = len([x for x in items if x.get('matched_any')])
        rate = (matched_count/total*100) if total>0 else 0.0
        # تعبئة الجدول
        self.table_kw.setRowCount(len(items))
        for i,it in enumerate(items):
            row_vals = [
                it.get("key",""), it.get("value_name",""), it.get("value_str",""), it.get("matched_kw",""),
                it.get("value_type",""), it.get("last_mod",""), it.get("owner",""), it.get("state",""),
                ", ".join(it.get("reasons",[]))
            ]
            for c,val in enumerate(row_vals):
                cell = QTableWidgetItem(str(val))
                cell.setTextAlignment(Qt.AlignLeft|Qt.AlignVCenter)
                if it.get("matched_any"):
                    cell.setBackground(QColor(35, 52, 93) if self.config.get("theme") in ("dark","midnight","ocean","steel","forest","ruby") else QColor(223, 230, 255))
                self.table_kw.setItem(i,c,cell)
        # إحصاءات
        self.lbl_total_kw.setText(str(total)); self.lbl_susp_kw.setText(str(matched_count))
        self.lbl_rate_kw.setText(f"{rate:.2f}%")
        self._update_status_counts(total, matched_count, rate)
        self.status.showMessage(tr("done").format(matched_count, total))
        if getattr(self, 'plot_kw', None):
            self._plot_reasons(self.plot_kw, items, which="kw")

    def _fill_table_and_stats_rules(self, items, total):
        matched_count = len([x for x in items if x.get('matched_any')])
        rate = (matched_count/total*100) if total>0 else 0.0
        self.table_rules.setRowCount(len(items))
        for i,it in enumerate(items):
            row = [
                it.get("key",""), it.get("value_name",""), it.get("value_str",""),
                it.get("value_type",""), it.get("last_mod",""), it.get("owner",""), it.get("state",""),
                it.get("matched_rule",""), ", ".join(it.get("reasons",[]))
            ]
            for c,val in enumerate(row):
                cell = QTableWidgetItem(str(val))
                cell.setTextAlignment(Qt.AlignLeft|Qt.AlignVCenter)
                if it.get("matched_any"):
                    cell.setBackground(QColor(35, 52, 93) if self.config.get("theme") in ("dark","midnight","ocean","steel","forest","ruby") else QColor(223, 230, 255))
                self.table_rules.setItem(i,c,cell)
        self.lbl_total_rules.setText(str(total)); self.lbl_susp_rules.setText(str(matched_count))
        self.lbl_rate_rules.setText(f"{rate:.2f}%")
        self._update_status_counts(total, matched_count, rate)
        self.status.showMessage(tr("done").format(matched_count, total))
        if getattr(self, 'plot_rules', None):
            self._plot_reasons(self.plot_rules, items, which="rules")

    def _update_status_counts(self, total:int, suspicious:int, rate:float):
        if LANG == "ar":
            self.lbl_status_total.setText(f"الإجمالي: {total}")
            self.lbl_status_susp.setText(f"المشبوه: {suspicious}")
            self.lbl_status_rate.setText(f"{rate:.2f}%")
        else:
            self.lbl_status_total.setText(f"Total: {total}")
            self.lbl_status_susp.setText(f"Suspicious: {suspicious}")
            self.lbl_status_rate.setText(f"{rate:.2f}%")

    # --- رسم تفاعلي وتحسينات ألوان وTooltips
    def _plot_reasons(self, plot, items, which: str):
        plot.clear()
        counts: Dict[str,int] = {}
        total = len(items) if items else 0
        for it in items:
            for r in it.get("reasons",[]):
                key = tr("reason_rule") if r.startswith(tr("reason_rule")) else r
                counts[key] = counts.get(key,0)+1
        if not counts:
            return
        xs = list(range(len(counts)))
        names = list(counts.keys())
        ys = [counts[k] for k in names]
        palette = [
            (58,134,255), (138,43,226), (255,99,132),
            (255,159,64), (75,192,192), (153,102,255),
            (255,205,86), (100,255,100), (0,200,180), (200,120,255)
        ]
        bars = []
        for i,(x,h) in enumerate(zip(xs,ys)):
            color = palette[i % len(palette)]
            try:
                bar = pg.BarGraphItem(x=[x], height=[h], width=0.7, brush=color, pen=(230,230,230))
                bar._reason_name = names[i]
                bar._count = h
                bar._ratio = (h/total*100) if total>0 else 0.0
                plot.addItem(bar)
                bars.append(bar)
            except Exception:
                pass
        try:
            ax = plot.getAxis('bottom')
            ax.setTicks([[(i, names[i]) for i in xs]])
            plot.showGrid(x=True, y=True, alpha=0.3)
        except Exception:
            pass

        # Tooltips + نقر لتطبيق فلترة
        def on_mouse_move(evt):
            pos = evt
            if hasattr(plot, 'vb'):
                vb = plot.getViewBox()
            else:
                vb = plot.plotItem.vb
            point = vb.mapSceneToView(evt)
            # تقريب x إلى أقرب عمود
            nearest = None
            min_dist = 1e9
            for i,x in enumerate(xs):
                dist = abs(point.x() - x)
                if dist < min_dist:
                    min_dist = dist
                    nearest = i
            if nearest is not None and 0 <= nearest < len(bars):
                b = bars[nearest]
                QToolTip_show = QApplication.instance()
                # عرض Tooltip يدوي عبر QToolTip
                tip = f"{b._reason_name}\nCount: {b._count}\nRate: {b._ratio:.2f}%"
                # لا توجد QToolTip سهلة هنا؛ يمكن استخدام setToolTip على الويجت
                plot.setToolTip(tip)

        def on_mouse_click(evt):
            if evt.button() == Qt.LeftButton:
                vb = plot.getViewBox() if hasattr(plot,'vb') else plot.plotItem.vb
                p = vb.mapSceneToView(evt.scenePos())
                # العثور على أقرب عمود
                nearest = None
                min_dist = 1e9
                for i,x in enumerate(xs):
                    dist = abs(p.x() - x)
                    if dist < min_dist:
                        min_dist = dist
                        nearest = i
                if nearest is not None and 0 <= nearest < len(bars):
                    reason_name = bars[nearest]._reason_name
                    # تطبيق فلترة على الجدول وفق السبب
                    self._apply_reason_filter(reason_name, which)

        # ربط الأحداث
        plot.scene().sigMouseMoved.connect(on_mouse_move)
        plot.scene().sigMouseClicked.connect(on_mouse_click)

    def _apply_reason_filter(self, reason_name: str, which: str):
        # يطبّق فلترة نصية على عمود "الأسباب" لاحتواء السبب
        cfg = {"column": None, "mode": "partial", "text": reason_name}
        if which == "kw":
            self._apply_table_filter_adv(self.table_kw, cfg)
        else:
            self._apply_table_filter_adv(self.table_rules, cfg)

    def _on_error(self, msg:str):
        QMessageBox.warning(self, tr("title"), msg)
        self.act_scan.setEnabled(True); self.act_stop.setEnabled(False); self.act_refresh.setEnabled(True)
        self.progress.setVisible(False); self.status.clearMessage()
        self.ui_heartbeat.stop()

    # ---------- فلترة + تفاصيل
    def _apply_table_filter_adv(self, table: QTableWidget, cfg: Dict[str, Any]):
        col = cfg.get("column", None)  # None يعني كل الأعمدة
        mode = cfg.get("mode", "partial")
        text = str(cfg.get("text",""))
        rows = table.rowCount()
        if not text:
            for r in range(rows):
                table.setRowHidden(r, False)
            return
        regex = None
        if mode == "regex":
            try:
                regex = re.compile(text)
            except Exception:
                regex = None
        for r in range(rows):
            show = False
            cols = range(table.columnCount()) if col is None else [col]
            for c in cols:
                it = table.item(r,c)
                if not it:
                    continue
                cell_text = it.text()
                if mode == "partial":
                    if text in cell_text:
                        show = True; break
                elif mode == "exact":
                    if text == cell_text:
                        show = True; break
                elif mode == "regex":
                    if regex and regex.search(cell_text):
                        show = True; break
            table.setRowHidden(r, not show)

    def _apply_filter_statusbar(self, cfg: Dict[str, Any]):
        if self.current_scan_tab == "kw":
            self._apply_table_filter_adv(self.table_kw, cfg)
        else:
            self._apply_table_filter_adv(self.table_rules, cfg)

    def _open_result_details_row(self, items: List[Dict[str,Any]], item: QTableWidgetItem):
        row = item.row()
        if row < 0 or row >= len(items):
            return
        dlg = ResultDetailsDialog(items[row], self)
        dlg.exec_()

    # ---------- قائمة سياقية على النتائج (تبويبية)
    def _show_table_context_menu(self, table: QTableWidget, items: List[Dict[str, Any]], pos, table_kind: str):
        row = table.currentRow()
        if row < 0 or row >= len(items):
            return
        it = items[row]
        menu = QMenu(self)
        act_del = menu.addAction(icon_for_action("delete"), tr("ctx_delete_value"))
        act_edit = menu.addAction(icon_for_action("edit"), tr("ctx_edit_value"))
        menu.addSeparator()
        act_go = menu.addAction(icon_for_action("go"), tr("ctx_go_to_key"))
        menu.addSeparator()
        act_copy_path = menu.addAction(icon_for_action("copy"), tr("ctx_copy_path"))
        act_copy_value = menu.addAction(icon_for_action("value"), tr("ctx_copy_value"))
        action = menu.exec_(table.viewport().mapToGlobal(pos))
        if not action:
            return
        if action == act_del:
            self._ctx_delete_value(it, items, table, row)
        elif action == act_edit:
            self._ctx_edit_value(it, items, table, row)
        elif action == act_go:
            self._ctx_go_to_key(it)
        elif action == act_copy_path:
            QApplication.clipboard().setText(it.get("key",""))
            QMessageBox.information(self, tr("title"), tr("action_done"))
        elif action == act_copy_value:
            QApplication.clipboard().setText(it.get("value_str",""))
            QMessageBox.information(self, tr("title"), tr("action_done"))

    def _ctx_delete_value(self, it: Dict[str, Any], items_list: List[Dict[str,Any]], table: QTableWidget, row_index: int):
        try:
            if not self._confirm(tr("confirm_delete_registry_value")):
                return
            hive, sub = parse_registry_path(it.get("key",""))
            if hive is None:
                raise Exception("Bad key")
            with winreg.OpenKey(hive, sub, 0, winreg.KEY_SET_VALUE) as k:
                winreg.DeleteValue(k, it.get("value_name",""))
            QMessageBox.information(self, tr("title"), tr("action_done"))
            table.removeRow(row_index)
            try:
                items_list.remove(it)
            except ValueError:
                pass
        except Exception as e:
            QMessageBox.warning(self, tr("title"), f"{tr('action_failed')}: {e}")

    def _ctx_edit_value(self, it: Dict[str, Any], items_list: List[Dict[str,Any]], table: QTableWidget, row_index: int):
        try:
            dlg = QDialog(self); dlg.setWindowTitle(tr("edit_value_title")); dlg.setWindowIcon(icon_for_action("edit")); dlg.resize(520, 180)
            v = QVBoxLayout(dlg)
            v.addWidget(QLabel(f"{tr('edit_value_new')}"))
            edit = QLineEdit(str(it.get("value_str",""))); v.addWidget(edit)
            btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            btns.button(QDialogButtonBox.Ok).setText(tr("ok"))
            btns.button(QDialogButtonBox.Cancel).setText(tr("cancel"))
            v.addWidget(btns)
            btns.accepted.connect(dlg.accept); btns.rejected.connect(dlg.reject)
            if dlg.exec_() != QDialog.Accepted:
                return
            new_val = edit.text()
            hive, sub = parse_registry_path(it.get("key",""))
            if hive is None:
                raise Exception("Bad key")
            vtype = it.get("value_type_raw", winreg.REG_SZ)
            write_val = new_val
            if vtype in (winreg.REG_DWORD, getattr(winreg, "REG_QWORD", 11)):
                write_val = int(new_val, 0) if isinstance(new_val, str) else int(new_val)
            elif vtype == winreg.REG_MULTI_SZ:
                write_val = [s.strip() for s in new_val.split(";")]
            elif vtype == winreg.REG_BINARY:
                write_val = bytes.fromhex(new_val.replace(" ", ""))
            with winreg.OpenKey(hive, sub, 0, winreg.KEY_SET_VALUE) as k:
                winreg.SetValueEx(k, it.get("value_name",""), 0, vtype, write_val)
            QMessageBox.information(self, tr("title"), tr("action_done"))
            it["value_str"] = new_val
            if self.current_scan_tab == "kw":
                if 0 <= row_index < self.table_kw.rowCount():
                    self.table_kw.item(row_index, 2).setText(new_val)
            else:
                if 0 <= row_index < self.table_rules.rowCount():
                    self.table_rules.item(row_index, 2).setText(new_val)
        except Exception as e:
            QMessageBox.warning(self, tr("title"), f"{tr('action_failed')}: {e}")

    def _ctx_go_to_key(self, it: Dict[str, Any]):
        dlg = RegistryBrowserDialog(self)
        path = it.get("key","")
        parts = path.split("\\")
        if parts:
            def find_child(parent_item, name):
                for i in range(parent_item.childCount()):
                    if parent_item.child(i).text(0) == name:
                        return parent_item.child(i)
                return None
            root_name = parts[0]
            root_item = None
            for i in range(dlg.tree.topLevelItemCount()):
                if dlg.tree.topLevelItem(i).text(0) == root_name:
                    root_item = dlg.tree.topLevelItem(i); break
            if root_item:
                dlg.tree.expandItem(root_item); dlg._on_expand(root_item)
                current = root_item
                for p in parts[1:]:
                    dlg.tree.expandItem(current); dlg._on_expand(current)
                    child = find_child(current, p)
                    if child:
                        current = child
                dlg.tree.setCurrentItem(current)
        dlg.exec_()

    # ---------- تصدير
    def _export(self):
        if self.current_scan_tab == "kw":
            data = self.last_kw
            if not data:
                QMessageBox.information(self, tr("title"), tr("no_results_to_export")); return
        else:
            data = self.last_rules_res
            if not data:
                QMessageBox.information(self, tr("title"), tr("no_results_to_export")); return

        btn = QMessageBox(self)
        btn.setWindowTitle(tr("export"))
        btn.setText(tr("export"))
        b_excel = btn.addButton(tr("export_excel"), QMessageBox.AcceptRole)
        b_html = btn.addButton(tr("export_html"), QMessageBox.ActionRole)
        btn.addButton(tr("cancel"), QMessageBox.RejectRole)
        btn.exec_()
        clicked = btn.clickedButton()
        if clicked == b_excel:
            self._export_excel_for(data)
        elif clicked == b_html:
            self._export_html_for(data)

    def _export_report_header(self) -> Dict[str, Any]:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if self.current_scan_tab == "kw":
            crit = self._criteria_keywords()
        else:
            crit = self._criteria_rules()
        owner_map_disp = {
            "all": tr("owner_all"),
            "systems": tr("owner_systems"),
            "localsystem": tr("owner_localsystem"),
            "users": tr("owner_users"),
        }
        header = {
            "title": tr("report_title"),
            "time_label": tr("report_time"),
            "time": now,
            "criteria_label": tr("report_criteria"),
            "criteria": {
                tr("criteria_keys"): ", ".join(crit.keys),
                tr("criteria_keywords"): ", ".join(crit.keywords),
                tr("criteria_value_type"): crit.value_type,
                tr("criteria_age"): str(crit.days if crit.use_age else 0),
                tr("criteria_accounts"): owner_map_disp.get(crit.owner_filter, tr("owner_all")),
            }
        }
        return header

    def _export_excel_for(self, data: List[Dict[str, Any]]):
        if openpyxl is None:
            QMessageBox.warning(self, "تنبيه" if LANG=="ar" else "Note", tr("need_openpyxl")); return
        fname, _ = QFileDialog.getSaveFileName(self, "حفظ Excel" if LANG=="ar" else "Save Excel", str(Path.home()), "Excel (*.xlsx)")
        if not fname: return
        try:
            header = self._export_report_header()
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "RegistryScan"
            # أنماط
            title_font = Font(size=14, bold=True)
            hdr_font = Font(bold=True)
            wrap = Alignment(wrap_text=True, vertical="top")
            th_fill = PatternFill("solid", fgColor="EEF2FF")
            border = Border(left=Side(style="thin", color="CCCCCC"),
                            right=Side(style="thin", color="CCCCCC"),
                            top=Side(style="thin", color="CCCCCC"),
                            bottom=Side(style="thin", color="CCCCCC"))

            ws.cell(row=1, column=1, value=header["title"]).font = title_font
            ws.cell(row=2, column=1, value=f"{header['time_label']}: {header['time']}")
            r = 3
            ws.cell(row=r, column=1, value=header["criteria_label"]).font = hdr_font; r += 1
            for k, v in header["criteria"].items():
                ws.cell(row=r, column=1, value=f"{k}: {v}"); r += 1
            r += 1

            # رؤوس الأعمدة حسب التبويب (لتطابق الجدول المعروض)
            if self.current_scan_tab == "kw":
                headers = [h for h in tr("tbl_headers") if h not in ([ "Matched rule" ] if LANG=="en" else ["القاعدة المطابقة"])]
            else:
                headers = [h for h in tr("tbl_headers") if h not in ([ "Matched keyword" ] if LANG=="en" else ["الكلمة المطابقة"])]

            for i,h in enumerate(headers, start=1):
                cell = ws.cell(row=r,column=i,value=h)
                cell.font = hdr_font
                cell.fill = th_fill
                cell.alignment = wrap
                cell.border = border

            # صفوف البيانات
            for rr, it in enumerate(data, start=r+1):
                if self.current_scan_tab == "kw":
                    row_vals = [
                        it.get("key",""), it.get("value_name",""), it.get("value_str",""), it.get("matched_kw",""),
                        it.get("value_type",""), it.get("last_mod",""), it.get("owner",""), it.get("state",""),
                        "; ".join(it.get("reasons",[]))
                    ]
                else:
                    row_vals = [
                        it.get("key",""), it.get("value_name",""), it.get("value_str",""),
                        it.get("value_type",""), it.get("last_mod",""), it.get("owner",""), it.get("state",""),
                        it.get("matched_rule",""), "; ".join(it.get("reasons",[]))
                    ]
                for c,v in enumerate(row_vals, start=1):
                    cell = ws.cell(row=rr,column=c,value=v)
                    cell.alignment = wrap
                    cell.border = border

            max_cols = len(headers)
            for i in range(1, max_cols+1):
                col = get_column_letter(i)
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in ws[col])
                ws.column_dimensions[col].width = min(max_len+4, 80)
            wb.save(fname)
            QMessageBox.information(self, tr("title"), f"{tr('saved')}: {fname}")
        except Exception as e:
            QMessageBox.warning(self, tr("title"), str(e))

    def _export_html_for(self, data: List[Dict[str, Any]]):
        fname, _ = QFileDialog.getSaveFileName(self, "حفظ HTML" if LANG=="ar" else "Save HTML", str(Path.home()), "HTML (*.html)")
        if not fname: return
        try:
            header = self._export_report_header()
            # رؤوس حسب التبويب
            if self.current_scan_tab == "kw":
                headers = [h for h in tr("tbl_headers") if h not in ([ "Matched rule" ] if LANG=="en" else ["القاعدة المطابقة"])]
                row_maker = lambda it: [
                    it.get("key",""), it.get("value_name",""), it.get("value_str",""), it.get("matched_kw",""),
                    it.get("value_type",""), it.get("last_mod",""), it.get("owner",""), it.get("state",""),
                    "; ".join(it.get("reasons",[]))
                ]
            else:
                headers = [h for h in tr("tbl_headers") if h not in ([ "Matched keyword" ] if LANG=="en" else ["الكلمة المطابقة"])]
                row_maker = lambda it: [
                    it.get("key",""), it.get("value_name",""), it.get("value_str",""),
                    it.get("value_type",""), it.get("last_mod",""), it.get("owner",""), it.get("state",""),
                    it.get("matched_rule",""), "; ".join(it.get("reasons",[]))
                ]

            head_html = "".join(f"<th>{html.escape(h)}</th>" for h in headers)
            rows_html = []
            for it in data:
                vals = row_maker(it)
                rows_html.append("<tr>" + "".join(f"<td>{html.escape(str(v))}</td>" for v in vals) + "</tr>")
            crit_list = "".join(f"<li><b>{html.escape(k)}:</b> {html.escape(v)}</li>" for k,v in header["criteria"].items())
            doc = f"""<!DOCTYPE html>
<html lang="{ 'ar' if LANG=='ar' else 'en' }">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{html.escape(tr('results'))}</title>
<style>
:root {{
  --bg: #ffffff;
  --txt: #222;
  --muted: #445;
  --th: #eef2ff;
  --border: #cfd7ee;
  --alt: #fafbff;
}}
body {{ font-family: Segoe UI, Tahoma, Arial, sans-serif; margin: 16px; color: var(--txt); background: var(--bg); }}
header.report {{ background: var(--th); padding: 12px 14px; border:1px solid var(--border); border-radius:10px; margin-bottom: 14px; }}
h1 {{ margin: 0 0 6px 0; font-size: 20px; color:#2d3c77; }}
small {{ color:var(--muted); }}
ul {{ margin:8px 0 0 20px; }}
table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
th, td {{ border: 1px solid #ccc; padding: 8px; vertical-align: top; word-wrap: break-word; }}
th {{ background: var(--th); text-align: left; }}
tr:nth-child(even) {{ background: var(--alt); }}
caption {{ text-align: { 'right' if LANG=='ar' else 'left' }; font-weight: 700; margin-bottom: 10px; }}
</style>
</head>
<body>
<header class="report">
  <h1>{html.escape(header['title'])}</h1>
  <div><small>{html.escape(header['time_label'])}: {html.escape(header['time'])}</small></div>
  <div><b>{html.escape(header['criteria_label'])}</b>
    <ul>{crit_list}</ul>
  </div>
</header>
<table>
<caption>{html.escape(tr('results'))}</caption>
<thead><tr>{head_html}</tr></thead>
<tbody>
{''.join(rows_html)}
</tbody>
</table>
</body>
</html>"""
            Path(fname).write_text(doc, encoding="utf-8")
            QMessageBox.information(self, tr("title"), f"{tr('saved')}: {fname}")
        except Exception as e:
            QMessageBox.warning(self, tr("title"), str(e))

    # ---------- التخزين للإعدادات
    def _load_config(self):
        global LANG
        if CONFIG_FILE.exists():
            try:
                cfg = json.load(open(CONFIG_FILE, "r", encoding="utf-8"))
                # ترقية الإعدادات القديمة: تحويل accounts -> owner_filter إن وجدت
                if "owner_filter" not in cfg:
                    if cfg.get("accounts"):
                        cfg["owner_filter"] = "users"
                    else:
                        cfg["owner_filter"] = "all"
                    cfg.pop("accounts", None)
                self.config.update(cfg)
            except Exception:
                pass
        LANG = self.config.get("lang","ar")

    def _save_config(self):
        try:
            cfg = dict(self.config)
            cfg.pop("accounts", None)
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    # ---------- تأكيدات
    def _confirm(self, message: str) -> bool:
        box = QMessageBox(self)
        box.setIcon(QMessageBox.Question)
        box.setWindowTitle(tr("confirm"))
        box.setText(message)
        ok_btn = box.addButton(tr("ok"), QMessageBox.AcceptRole)
        box.addButton(tr("cancel"), QMessageBox.RejectRole)
        box.exec_()
        return box.clickedButton() == ok_btn

# ================ تشغيل ================
def main():
    global LANG
    if CONFIG_FILE.exists():
        try:
            cfg = json.load(open(CONFIG_FILE, "r", encoding="utf-8"))
            LANG = cfg.get("lang","ar")
        except Exception:
            LANG = "ar"
    else:
        LANG = "ar"

    app = QApplication(sys.argv)
    w = Main()
    w.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
