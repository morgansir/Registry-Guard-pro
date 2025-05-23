# !/usr/bin/env python3
"""
تم تطوير هذا الكود كأداة مراقبة متقدمة للكشف عن الأنشطة الخبيثة،
مع واجهة مستخدم حديثة تحتوي على تأثيرات ألوان ثلاثية الأبعاد وعناصر بصرية جذابة،
بالإضافة إلى تقارير ورسوم بيانية تعرض تكرار النشاط ومدة النشاط.
تمت إضافة دعم استيراد قواعد SIGMA الرسمية (YAML)، حيث يمكن للمستخدم في تبويب Advanced Scan
تحميل ملف أو مجلد يحتوي على قواعد .yml وتحويلها إلى بنية داخلية وحفظها في قاعدة بيانات SQLite.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext, simpledialog
import sqlite3, hashlib, os, platform, ctypes, threading, time, logging, glob, subprocess, locale, json, hmac, \
    traceback, getpass
from datetime import datetime
from PIL import Image, ImageTk, ImageDraw
import numpy as np
import pandas as pd
from sklearn.ensemble import IsolationForest
import concurrent.futures
import matplotlib

matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt

try:
    import winreg  # Windows-only
except ImportError:
    winreg = None

from openpyxl.utils import get_column_letter  # لإصلاح عملية تصدير Excel

# تم إضافة الاستيراد من مكتبة PyYAML لدعم قواعد SIGMA
try:
    import yaml
except ImportError:
    messagebox.showerror("Dependency Error", "يرجى تثبيت مكتبة PyYAML باستخدام الأمر 'pip install pyyaml'")
    raise

# إعدادات التسجيل
logging.basicConfig(filename='monitor_secure.log', level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(message)s')
error_handler = logging.FileHandler('monitor_secure_errors.log')
error_handler.setLevel(logging.ERROR)
error_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
error_handler.setFormatter(error_formatter)
logging.getLogger().addHandler(error_handler)

try:
    locale.setlocale(locale.LC_COLLATE, 'en_US.UTF-8')
except Exception as e:

    logging.error("Locale setting error: " + str(e))
    locale.setlocale(locale.LC_COLLATE, '')

# ------------------- Constants and Colors ---------------------
PRIMARY_COLOR = "#4F46E5"  # أزرق نيلي
SECONDARY_COLOR = "#3B82F6"  # أزرق سماوي
TERTIARY_COLOR = "#06B6D4"  # سماوي فاتح
ACCENT_COLOR = "#10B981"  # أخضر
WARNING_COLOR = "#F59E0B"  # برتقالي
DANGER_COLOR = "#EF4444"  # أحمر

WELCOME_TAB_BG = "#1E3A8A"
REALTIME_TAB_BG = "#2563EB"
FULL_TAB_BG = PRIMARY_COLOR
ADVANCED_TAB_BG = SECONDARY_COLOR
CUSTOM_TAB_BG = TERTIARY_COLOR
SUSPICIOUS_TAB_BG = DANGER_COLOR

BUTTON_BG = "#007acc"
BUTTON_HOVER_BG = "#005f99"

FONT_NAME = "Helvetica"
BOLD_FONT = (FONT_NAME, 12, "bold")
HEADER_FONT = (FONT_NAME, 28, "bold")
TAB_FONT = (FONT_NAME, 14, "bold")

DB_FILE = "registry_monitor_secure.db"
HMAC_SECRET_KEY = b"SuperSecretKeyForHMAC"

DEFAULT_SUSPICIOUS_WORDS = [
    'virus', 'hack', 'malware', 'exploit', 'trojan', 'keylogger', 'ransomware', 'spyware', 'rootkit'
]

CRITICAL_RISK_THRESHOLD = 80
SUSPICIOUS_RISK_THRESHOLD = 60
EXEC_COUNT_THRESHOLD = 3

EXCLUDED_ENTRIES = [
    "c:/windows/system32",
    "/etc/systemd/system",
]


# ------------------- Helper Functions ---------------------
def log_exception(message):
    logging.error(message + "\n" + traceback.format_exc())


def should_exclude(entry):
    path = entry.get('path', '').lower()
    name = entry.get('entry_name', '').lower()
    for ex in EXCLUDED_ENTRIES:
        if ex in path or ex in name:
            return True
    return False


def send_critical_alert(message):
    logging.critical("Critical Alert: " + message)
    # يمكن إضافة منطق إرسال البريد الإلكتروني هنا


def initialize_anomaly_detector():
    np.random.seed(42)
    X = np.random.rand(200, 3) * np.array([100, 10, 5])
    model = IsolationForest(contamination=0.1, random_state=42)
    model.fit(X)
    return model


anomaly_detector = initialize_anomaly_detector()


def calculate_risk_advanced(content, execution_count=1):
    try:
        text = str(content).lower()
        risk_score = 0
        for word in DEFAULT_SUSPICIOUS_WORDS:
            count = text.count(word.lower())
            risk_score += count * 20
        risk_score += (execution_count - 1) * 5
        risk_score = min(risk_score, 100)
        return f"{risk_score}%"
    except Exception as e:
        log_exception("Advanced risk calculation error: " + str(e))
        return "0%"


def verify_digital_signature(content, provided_signature):
    try:
        computed_hmac = hmac.new(HMAC_SECRET_KEY, str(content).encode(), hashlib.sha256).hexdigest()[:12]
        valid = hmac.compare_digest(computed_hmac, provided_signature)
        return valid
    except Exception as e:
        log_exception("Digital signature verification error: " + str(e))
        return False


def ml_anomaly_score(entry):
    try:
        risk_str = entry.get('risk_percentage', '0%')
        risk_num = int(risk_str.replace('%', ''))
        exec_count = entry.get('execution_count', 1)
        score = (risk_num / 100) * (exec_count / 5)
        return score
    except Exception as e:
        log_exception("ML anomaly score error: " + str(e))
        return 0


def advanced_risk_assessment(entry):
    if should_exclude(entry):
        return "Normal"
    try:
        risk_str = entry.get('risk_percentage', '0%')
        risk_num = int(risk_str.replace('%', ''))
    except:
        risk_num = 0
    exec_count = entry.get('execution_count', 1)
    provided_signature = entry.get('digital_signature', '')
    digital_valid = verify_digital_signature(entry.get('hash', ''), provided_signature)
    anomaly_score = ml_anomaly_score(entry)
    source_lower = entry.get('source', '').lower()
    if source_lower in ['windows registry', 'registry custom scan', 'windows event log', 'linux cron', 'linux log',
                        'systemd service'] and risk_num < 50:
        return "Normal"
    if exec_count > 1 and risk_num <= 30:
        return "Normal"
    features = [[risk_num, exec_count, anomaly_score * 100]]
    try:
        prediction = anomaly_detector.predict(features)[0]
    except Exception as e:
        log_exception("Prediction error: " + str(e))
        prediction = 1
    if prediction == -1:
        return "Suspicious"
    if risk_num >= CRITICAL_RISK_THRESHOLD or anomaly_score >= 0.9 or (not digital_valid and risk_num >= 50):
        return "Critical"
    elif risk_num >= SUSPICIOUS_RISK_THRESHOLD or anomaly_score >= 0.7 or (
            exec_count >= EXEC_COUNT_THRESHOLD and not digital_valid):
        return "Suspicious"
    return entry.get('severity', 'Normal')


def compute_frequency(entry):
    try:
        timestamp_str = entry.get('timestamp', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
        days = max((datetime.now() - timestamp).days, 1)
        freq = entry.get('execution_count', 1) / days
        if freq >= 1:
            return f"{freq:.1f} times/day"
        elif freq * 7 >= 1:
            return f"{(freq * 7):.1f} times/week"
        else:
            return f"{(freq * 30):.1f} times/month"
    except Exception as e:
        log_exception("Frequency computation error: " + str(e))
        return "N/A"


def compute_activity_age(entry):
    try:
        timestamp_str = entry.get('timestamp', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
        days = (datetime.now() - timestamp).days
        return f"{days} day(s)"
    except Exception as e:
        log_exception("Activity age computation error: " + str(e))
        return "N/A"


# ------------------- Database Creation and Alterations ---------------------
def create_database():
    try:
        with sqlite3.connect(DB_FILE) as conn:
            c = conn.cursor()
            # إنشاء جدول الإدخالات
            c.execute('''
                CREATE TABLE IF NOT EXISTS registry_entries(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source TEXT,
                    entry_name TEXT,
                    path TEXT,
                    severity TEXT,
                    timestamp TEXT,
                    hash TEXT,
                    system_type TEXT,
                    current_user TEXT,
                    execution_count INTEGER,
                    digital_signature TEXT,
                    risk_percentage TEXT,
                    frequency TEXT,
                    activity_age TEXT
                )
            ''')
            c.execute("PRAGMA table_info(registry_entries)")
            existing_columns = [col[1] for col in c.fetchall()]
            new_columns = [
                ('current_user', 'TEXT'),
                ('execution_count', 'INTEGER'),
                ('digital_signature', 'TEXT'),
                ('risk_percentage', 'TEXT'),
                ('frequency', 'TEXT'),
                ('activity_age', 'TEXT')
            ]
            for col_name, col_type in new_columns:
                if col_name not in existing_columns:
                    try:
                        c.execute(f"ALTER TABLE registry_entries ADD COLUMN {col_name} {col_type}")
                    except Exception as e:
                        log_exception(f"Error adding column {col_name}: {str(e)}")
            # إنشاء جدول لقواعد SIGMA
            c.execute('''

                CREATE TABLE IF NOT EXISTS sigma_rules(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    filename TEXT,
                    rule_data TEXT,
                    imported_at TEXT
                )
            ''')
            conn.commit()
    except Exception as e:
        log_exception("Database creation error: " + str(e))


# ------------------- Helper for إنشاء صورة متدرجة للخلفيات ---------------------
def create_gradient_image(width, height, color1, color2):
    base = Image.new('RGB', (width, height), color1)
    top = Image.new('RGB', (width, height), color2)
    mask = Image.new('L', (width, height))
    mask_data = []
    for y in range(height):
        mask_data.extend([int(255 * (y / height))] * width)
    mask.putdata(mask_data)
    base.paste(top, (0, 0), mask)
    return ImageTk.PhotoImage(base)


# ------------------- System Scanning Functions ---------------------
def windows_scan_all():
    entries = []
    try:
        current_user = getpass.getuser()
    except Exception as e:
        log_exception("Error getting current user: " + str(e))
        current_user = "Unknown"

    def scan_target_key(hive, key_path):
        if winreg is None:
            return
        try:
            with winreg.OpenKey(hive, key_path, 0, winreg.KEY_READ) as key:
                num_values = winreg.QueryInfoKey(key)[1]
                for i in range(num_values):
                    try:
                        name, value, _ = winreg.EnumValue(key, i)
                        risk = calculate_risk_advanced(str(value), execution_count=1)
                        entry = {
                            'source': 'Windows Registry',
                            'entry_name': name,
                            'path': key_path,
                            'severity': 'Normal',
                            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            'hash': hashlib.sha256(str(value).encode()).hexdigest(),
                            'system_type': 'Windows',
                            'current_user': current_user,

                            'execution_count': 1,
                            'digital_signature': hmac.new(HMAC_SECRET_KEY,
                                                          hashlib.sha256(str(value).encode()).hexdigest().encode(),
                                                          hashlib.sha256).hexdigest()[:12],
                            'risk_percentage': risk
                        }
                        if should_exclude(entry):
                            continue
                        entry['severity'] = advanced_risk_assessment(entry)
                        entry['frequency'] = compute_frequency(entry)
                        entry['activity_age'] = compute_activity_age(entry)
                        entries.append(entry)
                    except Exception as ve:
                        log_exception(f"Error reading value in {key_path}: " + str(ve))
        except Exception as e:
            log_exception(f"Error opening key {key_path}: " + str(e))

    target_keys = [
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Run"),
        (winreg.HKEY_CURRENT_USER, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Run"),
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows\CurrentVersion\RunOnce"),
        (winreg.HKEY_CURRENT_USER, r"SOFTWARE\Microsoft\Windows\CurrentVersion\RunOnce"),
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows NT\CurrentVersion\Winlogon"),
    ]
    for hive, key_path in target_keys:
        scan_target_key(hive, key_path)
    return entries


def linux_scan():
    entries = []
    current_user = os.getenv('USER', 'root')
    try:
        cron_files = ['/etc/crontab', '/etc/cron.d/', '/var/spool/cron/crontabs/']
        for pattern in cron_files:
            for file in glob.glob(pattern):
                try:
                    with open(file, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                    risk = calculate_risk_advanced(content, execution_count=1)
                    entry = {
                        'source': 'Linux Cron',
                        'entry_name': os.path.basename(file),
                        'path': file,
                        'severity': 'Normal',
                        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        'hash': hashlib.sha256(content.encode()).hexdigest(),
                        'system_type': 'Linux',
                        'current_user': current_user,
                        'execution_count': 1,
                        'digital_signature': hmac.new(HMAC_SECRET_KEY,
                                                      hashlib.sha256(content.encode()).hexdigest().encode(),

                                                      hashlib.sha256).hexdigest()[:12],
                        'risk_percentage': risk
                    }
                    if should_exclude(entry):
                        continue
                    entry['severity'] = advanced_risk_assessment(entry)
                    entry['frequency'] = compute_frequency(entry)
                    entry['activity_age'] = compute_activity_age(entry)
                    entries.append(entry)
                except Exception as ex:
                    log_exception(f"Error reading file {file}: " + str(ex))
        services = subprocess.run(['systemctl', 'list-unit-files', '--type=service'],
                                  stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True).stdout.split('\n')
        for service in services[1:-6]:
            parts = service.split()
            if len(parts) >= 2 and parts[1] == 'enabled':
                entry = {
                    'source': 'Systemd Service',
                    'entry_name': parts[0],
                    'path': f"/etc/systemd/system/{parts[0]}",
                    'severity': 'Normal',
                    'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'hash': '',
                    'system_type': 'Linux',
                    'current_user': current_user,
                    'execution_count': 1,
                    'digital_signature': '',
                    'risk_percentage': "0%"
                }
                if should_exclude(entry):
                    continue
                entry['severity'] = advanced_risk_assessment(entry)
                entry['frequency'] = compute_frequency(entry)
                entry['activity_age'] = compute_activity_age(entry)
                entries.append(entry)
    except Exception as e:
        log_exception("Linux Scan Error: " + str(e))
    return entries


def linux_log_scan():
    entries = []
    current_user = os.getenv('USER', 'root')
    log_files = ['/var/log/syslog', '/var/log/auth.log', '/var/log/messages', '/var/log/dmesg']
    for log_file in log_files:
        if os.path.exists(log_file):
            try:
                with open(log_file, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                risk = calculate_risk_advanced(content, execution_count=1)
                entry = {
                    'source': 'Linux Log',

                    'entry_name': os.path.basename(log_file),
                    'path': log_file,
                    'severity': 'Normal',
                    'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'hash': hashlib.sha256(content.encode()).hexdigest(),
                    'system_type': 'Linux',
                    'current_user': current_user,
                    'execution_count': 1,
                    'digital_signature': hmac.new(HMAC_SECRET_KEY,
                                                  hashlib.sha256(content.encode()).hexdigest().encode(),
                                                  hashlib.sha256).hexdigest()[:12],
                    'risk_percentage': risk
                }
                if should_exclude(entry):
                    continue
                entry['severity'] = advanced_risk_assessment(entry)
                entry['frequency'] = compute_frequency(entry)
                entry['activity_age'] = compute_activity_age(entry)
                entries.append(entry)
            except Exception as ex:
                log_exception(f"Error reading log file {log_file}: " + str(ex))
    return entries


def windows_event_log_scan():
    entries = []
    try:
        import win32evtlog
    except ImportError:
        win32evtlog = None
    if win32evtlog is None:
        return entries
    try:
        server = 'localhost'
        for logtype in ['Application', 'System']:
            hand = win32evtlog.OpenEventLog(server, logtype)
            flags = win32evtlog.EVENTLOG_BACKWARDS_READ | win32evtlog.EVENTLOG_SEQUENTIAL_READ
            events = win32evtlog.ReadEventLog(hand, flags, 0)
            if events:
                for event in events:
                    description = event.StringInserts if event.StringInserts else []
                    description_str = " ".join(description) if description else ""
                    risk = calculate_risk_advanced(description_str, execution_count=1)
                    entry = {
                        'source': 'Windows Event Log',
                        'entry_name': logtype,
                        'path': f"Event ID: {event.EventID}",
                        'severity': 'Normal',
                        'timestamp': event.TimeGenerated.Format() if hasattr(event.TimeGenerated,
                                                                             'Format') else datetime.now().strftime(
                            "%Y-%m-%d %H:%M:%S"),
                        'hash': hashlib.sha256(description_str.encode()).hexdigest(),
                        'system_type': 'Windows',

                        'current_user': getpass.getuser(),
                        'execution_count': 1,
                        'digital_signature': hmac.new(HMAC_SECRET_KEY,
                                                      hashlib.sha256(description_str.encode()).hexdigest().encode(),
                                                      hashlib.sha256).hexdigest()[:12],
                        'risk_percentage': risk
                    }
                    if should_exclude(entry):
                        continue
                    entry['severity'] = advanced_risk_assessment(entry)
                    entry['frequency'] = compute_frequency(entry)
                    entry['activity_age'] = compute_activity_age(entry)
                    entries.append(entry)
            win32evtlog.CloseEventLog(hand)
    except Exception as e:
        log_exception("Windows event log scan error: " + str(e))
    return entries


def all_system_logs_scan_optimized(suppress_message=False, update_existing=False, progress_callback=None):
    start_time = time.perf_counter()
    current_system = platform.system()
    entries = []
    tasks = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        if current_system == 'Windows':
            tasks.append(executor.submit(windows_scan_all))
            tasks.append(executor.submit(windows_event_log_scan))
        else:
            tasks.append(executor.submit(linux_scan))
            tasks.append(executor.submit(linux_log_scan))
        total_tasks = len(tasks)
        for i, future in enumerate(concurrent.futures.as_completed(tasks)):
            try:
                result = future.result()
                if result:
                    filtered_result = [entry for entry in result if not should_exclude(entry)]
                    entries.extend(filtered_result)
            except Exception as e:
                log_exception("Error in scanning thread: " + str(e))
            if progress_callback:
                progress = int(((i + 1) / total_tasks) * 100)
                progress_callback(progress)
    end_time = time.perf_counter()
    logging.info(
        f"All system logs scan completed in {end_time - start_time:.2f} seconds. Processed {len(entries)} records.")
    try:
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            insert_data = []
            for entry in entries:

                if update_existing:
                    cursor.execute(
                        "SELECT id, execution_count FROM registry_entries WHERE source=? AND entry_name=? AND path=?",
                        (entry['source'], entry['entry_name'], entry['path']))
                    result = cursor.fetchone()
                    if result:
                        record_id, exec_count = result
                        new_exec_count = (exec_count + 1) if exec_count is not None else 1
                        new_severity = advanced_risk_assessment({
                            'severity': entry['severity'],
                            'risk_percentage': entry['risk_percentage'],
                            'execution_count': new_exec_count,
                            'digital_signature': entry['digital_signature'],
                            'path': entry['path'],
                            'hash': entry['hash']
                        })
                        frequency = compute_frequency(entry)
                        age = compute_activity_age(entry)
                        cursor.execute('''
                            UPDATE registry_entries
                            SET severity=?, timestamp=?, hash=?, system_type=?, current_user=?, execution_count=?, digital_signature=?, risk_percentage=?, frequency=?, activity_age=?
                            WHERE id=?
                        ''', (
                        new_severity, entry['timestamp'], entry['hash'], entry['system_type'], entry['current_user'],
                        new_exec_count, entry['digital_signature'], entry['risk_percentage'], frequency, age,
                        record_id))
                        continue
                new_severity = advanced_risk_assessment(entry)
                frequency = compute_frequency(entry)
                age = compute_activity_age(entry)
                insert_data.append((entry['source'], entry['entry_name'], entry['path'], new_severity,
                                    entry['timestamp'], entry['hash'], entry['system_type'], entry['current_user'],
                                    entry['execution_count'], entry['digital_signature'], entry['risk_percentage'],
                                    frequency, age))
            if insert_data:
                cursor.executemany('''
                    INSERT INTO registry_entries 
                    (source, entry_name, path, severity, timestamp, hash, system_type,
                     current_user, execution_count, digital_signature, risk_percentage, frequency, activity_age)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                ''', insert_data)
            conn.commit()
            logging.info("All system logs scan completed: Processed {} records".format(len(entries)))
            if not suppress_message:
                messagebox.showinfo("Success", f"Processed {len(entries)} records")
    except Exception as e:
        messagebox.showerror("Error", f"Scan failed: {str(e)}")
        log_exception("All system logs scan error: " + str(e))


# ------------------- Admin Privilege Functions ---------------------

def check_admin_privileges():
    if platform.system() == 'Windows':
        try:
            return ctypes.windll.shell32.IsUserAnAdmin() != 0
        except Exception as e:
            log_exception("Admin check error: " + str(e))
            return False
    else:
        return os.geteuid() == 0


# ------------------- Registry Selection Dialog ---------------------
REG_SELECTION_DATA = {}


def open_registry_tree_selection_dialog():
    selected_data = None
    dlg = tk.Toplevel()
    dlg.title("Select Registry Entry")
    dlg.geometry("600x400")
    pick_key_var = tk.BooleanVar(value=False)
    chk_pick_key = tk.Checkbutton(dlg, text="Select Entire Key", variable=pick_key_var, font=BOLD_FONT)
    chk_pick_key.pack(anchor='w', padx=5, pady=5)
    tree = ttk.Treeview(dlg)
    tree.heading("#0", text="Registry Tree", anchor="w")
    tree.pack(fill="both", expand=True, side=tk.LEFT)
    scrollbar = ttk.Scrollbar(dlg, orient="vertical", command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    scrollbar.pack(fill="y", side=tk.RIGHT)
    hives = [
        ("HKEY_LOCAL_MACHINE", None),
        ("HKEY_CLASSES_ROOT", None),
        ("HKEY_USERS", None),
        ("HKEY_CURRENT_USER", None),
        ("HKEY_CURRENT_CONFIG", None)
    ]
    for hive_name, _ in hives:
        node = tree.insert("", "end", text=hive_name, values=("key", hive_name))
        REG_SELECTION_DATA[node] = {"type": "key", "key_path": hive_name}
        tree.insert(node, 'end', text="dummy")

    def populate_tree(parent, key_path):
        try:
            i = 0
            parts = key_path.split("\\", 1)
            hive_str = parts[0]
            sub_key = parts[1] if len(parts) > 1 else ""
            hive = getattr(winreg, hive_str, None)
            if hive is None:
                return
            key_handle = winreg.OpenKey(hive, sub_key) if sub_key else winreg.OpenKey(hive, "")
            while True:
                subkey = winreg.EnumKey(key_handle, i)

                full_path = key_path + "\\" + subkey if key_path else subkey
                node_id = tree.insert(parent, 'end', text=subkey, values=("key", full_path))
                REG_SELECTION_DATA[node_id] = {"type": "key", "key_path": full_path}
                tree.insert(node_id, 'end', text="dummy")
                i += 1
        except Exception:
            pass
        try:
            j = 0
            key_handle = winreg.OpenKey(hive, sub_key) if sub_key else winreg.OpenKey(hive, "")
            while True:
                value_name, value_data, _ = winreg.EnumValue(key_handle, j)
                text = f"Value: {value_name} = {value_data}"
                node_id = tree.insert(parent, 'end', text=text, values=("value", key_path, value_name, value_data))
                REG_SELECTION_DATA[node_id] = {"type": "value", "key_path": key_path, "value_name": value_name,
                                               "value": value_data}
                j += 1
        except Exception:
            pass

    def on_open(event):
        node = tree.focus()
        if not node:
            return
        children = tree.get_children(node)
        if children and tree.item(children[0], "text") == "dummy":
            tree.delete(children[0])
            key_info = REG_SELECTION_DATA.get(node)
            if key_info and key_info.get("type") == "key":
                populate_tree(node, key_info.get("key_path"))

    tree.bind("<<TreeviewOpen>>", on_open)

    def confirm_selection():
        nonlocal selected_data
        sel = tree.selection()
        if not sel:
            messagebox.showwarning("Warning", "Please select an entry")
            return
        info = REG_SELECTION_DATA.get(sel[0])
        if not info:
            messagebox.showwarning("Warning", "Invalid selection")
            return
        if not pick_key_var.get() and info.get("type") != "value":
            messagebox.showwarning("Warning", "Please select a value entry or enable selecting the entire key")
            return
        selected_data = info
        dlg.destroy()

    btn_confirm = tk.Button(dlg, text="Confirm", command=confirm_selection, font=BOLD_FONT, bg=BUTTON_BG, fg="white",
                            relief="raised", bd=3)
    btn_confirm.pack(pady=5)
    dlg.wait_window()
    return selected_data


# ------------------- New Function: Block Record ---------------------
def block_record(source, entry_name, path_info):
    try:
        if source == "Windows Registry":
            if winreg is None:
                messagebox.showerror("Error", "winreg module not available")
                return False
            if not check_admin_privileges():
                messagebox.showerror("Error", "Insufficient privileges to block registry entry")
                return False
            if "HKEY_LOCAL_MACHINE" in path_info:
                hive = winreg.HKEY_LOCAL_MACHINE
            elif "HKEY_CURRENT_USER" in path_info:
                hive = winreg.HKEY_CURRENT_USER
            else:
                messagebox.showerror("Error", "Unknown registry hive")
                return False
            key_path = path_info.split(" (from")[0]
            with winreg.OpenKey(hive, key_path, 0, winreg.KEY_SET_VALUE) as reg_key:
                winreg.SetValueEx(reg_key, entry_name, 0, winreg.REG_SZ, "BLOCKED")
            logging.info(f"Registry entry {entry_name} blocked in {key_path}")
            return True
        elif source == "Linux Cron":
            if os.path.isfile(path_info):
                with open(path_info, 'r+', encoding='utf-8') as f:
                    content = f.read()
                    f.seek(0, 0)
                    f.write("# BLOCKED\n" + content)
                subprocess.run(['systemctl', 'reload', 'cron.service'], check=True)
                logging.info(f"Cron file {path_info} blocked")
                return True
            else:
                messagebox.showerror("Error", "Cron file not found")
                return False
        elif source == "Systemd Service":
            service_name = os.path.basename(path_info)
            subprocess.run(['systemctl', 'disable', service_name], check=True)
            logging.info(f"Service {service_name} disabled (blocked)")
            return True
        else:
            messagebox.showerror("Error", "Block operation not supported for this source")
            return False
    except Exception as e:
        messagebox.showerror("Error", f"Block failed: {str(e)}")
        log_exception("Block record error: " + str(e))

        return False


# ------------------- New Function: Import SIGMA Rules ---------------------
def import_sigma_rules_from_files(file_paths):
    try:
        imported_count = 0
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            for file_path in file_paths:
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        rule_content = f.read()
                    # استخدام PyYAML لقراءة الملف إلى بنية داخلية (dict)
                    rule_dict = yaml.safe_load(rule_content)
                    # الحصول على اسم القاعدة من المفتاح "title" إن وُجد، وإلا استخدام اسم الملف
                    rule_name = rule_dict.get("title", os.path.basename(file_path)) if isinstance(rule_dict,
                                                                                                  dict) else os.path.basename(
                        file_path)
                    # حفظ القاعدة في قاعدة البيانات بتنسيق JSON للنظام الداخلي
                    cursor.execute('''
                        INSERT INTO sigma_rules (filename, rule_data, imported_at)
                        VALUES (?, ?, ?)
                    ''', (
                    rule_name, json.dumps(rule_dict, ensure_ascii=False), datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
                    imported_count += 1
                except Exception as inner_e:
                    log_exception(f"Error importing rule from {file_path}: " + str(inner_e))
            conn.commit()
        return imported_count
    except Exception as e:
        log_exception("Import SIGMA rules error: " + str(e))
        return 0


# ------------------- Main Application and Screens ---------------------
class MainApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Advanced Monitoring System - Enhanced Edition")
        self.state("zoomed")
        self.configure(bg=PRIMARY_COLOR)
        try:
            self.iconbitmap("modern_logo.ico")
        except Exception as e:
            logging.warning("Icon file not found: " + str(e))
        create_database()

        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TNotebook", background="#f0f0f0", borderwidth=0)
        style.configure("TNotebook.Tab", padding=[10, 5], background="#cccccc", foreground="black")
        style.map("TNotebook.Tab", background=[("selected", "#007acc")])

        self.create_toolbar()
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=5, pady=5)

        self.tab_welcome = WelcomeScreen(self.notebook, self)
        self.notebook.add(self.tab_welcome, text="Welcome")
        self.tab_realtime = ScannerScreen(self.notebook, self, "realtime")
        self.notebook.add(self.tab_realtime, text="Real-Time Scan")
        self.tab_full = ScannerScreen(self.notebook, self, "full")
        self.notebook.add(self.tab_full, text="Full Scan")
        self.tab_advanced = ScannerScreen(self.notebook, self, "suspicious")
        self.notebook.add(self.tab_advanced, text="Advanced Scan")
        self.tab_custom = ScannerScreen(self.notebook, self, "custom")
        self.notebook.add(self.tab_custom, text="Custom Scan")
        self.tab_suspicious = SuspiciousScreen(self.notebook, self)
        self.notebook.add(self.tab_suspicious, text="Suspicious Logs")
        self.event_log = scrolledtext.ScrolledText(self, height=8, state='disabled', font=BOLD_FONT)
        self.event_log.pack(fill=tk.X, padx=10, pady=5)
        self.create_chart()

    def create_toolbar(self):
        toolbar = tk.Frame(self, bg=PRIMARY_COLOR)
        toolbar.pack(fill=tk.X)
        btn_exit = tk.Button(toolbar, text="Exit", font=BOLD_FONT, bg=BUTTON_BG, fg="white",
                             command=self.destroy, relief="raised", bd=3)
        btn_exit.pack(side=tk.RIGHT, padx=10, pady=5)

    def create_chart(self):
        fig, ax = plt.subplots(figsize=(5, 3), dpi=100)
        labels = ['Normal', 'Suspicious', 'Critical']
        counts = [10, 5, 2]
        ax.bar(labels, counts, color=[ACCENT_COLOR, WARNING_COLOR, DANGER_COLOR])
        ax.set_title("Record Distribution")
        self.chart_canvas = FigureCanvasTkAgg(fig, master=self)
        self.chart_canvas.draw()
        self.chart_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=False, padx=10, pady=5)

    def log_event(self, message):
        self.event_log.config(state='normal')
        timestamped_message = f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {message}\n"
        self.event_log.insert(tk.END, timestamped_message)
        self.event_log.see(tk.END)
        self.event_log.config(state='disabled')
        logging.info(message)


# شاشة الترحيب مع خلفية متدرجة ثلاثية الأبعاد
class WelcomeScreen(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg=WELCOME_TAB_BG)
        self.controller = controller
        self.canvas = tk.Canvas(self, bg=WELCOME_TAB_BG, highlightthickness=0)

        self.canvas.pack(fill="both", expand=True)
        self.circles = []
        for i in range(10):
            x, y = 50 * i, 50 * i
            r = 30
            circle = self.canvas.create_oval(x, y, x + r, y + r, fill=ACCENT_COLOR, outline="")
            self.circles.append((circle, 1, 1))
        self.title_label = tk.Label(self.canvas, text="Advanced Monitoring System\n", font=HEADER_FONT,
                                    bg=WELCOME_TAB_BG, fg="white")
        self.desc_label = tk.Label(self.canvas, text=("Welcome to the Advanced Monitoring System.\n\n"
                                                      "Features:\n"
                                                      "• Real-Time Scan\n"
                                                      "• Full System Scan\n"
                                                      "• Intelligent Advanced Detection\n"
                                                      "• Custom Registry Scan\n"
                                                      "• Interactive Reports and Charts\n"
                                                      "• Import SIGMA Rules via YAML"),
                                   font=(FONT_NAME, 16, "bold"), bg=WELCOME_TAB_BG, fg="white", justify="center")
        self.canvas.create_window(self.winfo_screenwidth() // 2, 100, window=self.title_label)
        self.canvas.create_window(self.winfo_screenwidth() // 2, 200, window=self.desc_label)
        self.animate_background()

    def animate_background(self):
        for index, (circle, dx, dy) in enumerate(self.circles):
            self.canvas.move(circle, dx, dy)
            x1, y1, x2, y2 = self.canvas.coords(circle)
            if x2 > self.winfo_width() or x1 < 0:
                dx = -dx
            if y2 > self.winfo_height() or y1 < 0:
                dy = -dy
            self.circles[index] = (circle, dx, dy)
        self.after(50, self.animate_background)


# شاشة الماسح (Scanner) مع تحسين التصميم
class ScannerScreen(tk.Frame):
    def __init__(self, parent, controller, scan_type):
        bg_color = {
            "realtime": REALTIME_TAB_BG,
            "full": FULL_TAB_BG,
            "suspicious": ADVANCED_TAB_BG,
            "custom": CUSTOM_TAB_BG
        }.get(scan_type, PRIMARY_COLOR)
        self.gradient_img = create_gradient_image(800, 600, bg_color, SECONDARY_COLOR)
        super().__init__(parent, bg=bg_color)
        self.controller = controller
        self.scan_type = scan_type
        header = tk.Label(self, text=f"Scan Interface: {scan_type.capitalize()}", font=TAB_FONT, bg=bg_color,
                          fg="white")
        header.pack(pady=10)
        if self.scan_type == "suspicious":
            config_frame = tk.LabelFrame(self, text="Suspicious Word Filter", font=BOLD_FONT, bg=bg_color,

                                         fg="white", bd=2, relief="ridge")
            config_frame.pack(padx=10, pady=10, fill=tk.X)
            options = ["All"] + DEFAULT_SUSPICIOUS_WORDS
            self.selected_suspicious = tk.StringVar(value="All")
            dropdown = ttk.OptionMenu(config_frame, self.selected_suspicious, options[0], *options)
            dropdown.config(style='TMenubutton')
            dropdown.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
            btn_delete_word = tk.Button(config_frame, text="Delete Word", font=BOLD_FONT, bg=BUTTON_BG, fg="white",
                                        command=self.delete_suspicious_word, relief="raised", bd=3)
            btn_delete_word.grid(row=0, column=2, padx=5, pady=5)
            btn_edit_word = tk.Button(config_frame, text="Edit Word", font=BOLD_FONT, bg=BUTTON_BG, fg="white",
                                      command=self.edit_suspicious_word, relief="raised", bd=3)
            btn_edit_word.grid(row=0, column=3, padx=5, pady=5)
            btn_add_word = tk.Button(config_frame, text="Add Word", font=BOLD_FONT, bg=BUTTON_BG, fg="white",
                                     command=self.add_suspicious_word, relief="raised", bd=3)
            btn_add_word.grid(row=0, column=4, padx=5, pady=5)
            # زر لاستيراد قواعد SIGMA (YAML)
            btn_import_sigma = tk.Button(config_frame, text="استيراد من ملف", font=BOLD_FONT, bg=BUTTON_BG, fg="white",
                                         command=self.import_sigma_rules, relief="raised", bd=3)
            btn_import_sigma.grid(row=0, column=5, padx=5, pady=5)
        tool_frame = tk.Frame(self, bg=bg_color)
        tool_frame.pack(fill=tk.X, padx=10, pady=5)
        self.btn_start = tk.Button(tool_frame, text="Start Scan", font=BOLD_FONT, bg=BUTTON_BG, fg="white",
                                   command=self.start_scan, relief="raised", bd=3)
        self.btn_start.pack(side=tk.LEFT, padx=5)
        if self.scan_type == "realtime":
            self.btn_stop = tk.Button(tool_frame, text="Stop Scan", font=BOLD_FONT, bg=BUTTON_BG, fg="white",
                                      command=self.stop_scan, relief="raised", bd=3)
            self.btn_stop.pack(side=tk.LEFT, padx=5)
        btn_refresh = tk.Button(tool_frame, text="Refresh Data", font=BOLD_FONT, bg=BUTTON_BG, fg="white",
                                command=self.refresh_data_thread, relief="raised", bd=3)
        btn_refresh.pack(side=tk.LEFT, padx=5)
        btn_del_all = tk.Button(tool_frame, text="Delete All Records", font=BOLD_FONT, bg=BUTTON_BG, fg="white",
                                command=self.delete_all_records, relief="raised", bd=3)
        btn_del_all.pack(side=tk.LEFT, padx=5)
        btn_block = tk.Button(tool_frame, text="Block Record", font=BOLD_FONT, bg=BUTTON_BG, fg="white",
                              command=self.block_selected, relief="raised", bd=3)
        btn_block.pack(side=tk.LEFT, padx=5)
        btn_export = tk.Button(tool_frame, text="Export Data", font=BOLD_FONT, bg=BUTTON_BG, fg="white",
                               command=self.export_data, relief="raised", bd=3)
        btn_export.pack(side=tk.LEFT, padx=5)
        self.progress_bar = ttk.Progressbar(self, orient="horizontal", mode="determinate", maximum=100)
        self.progress_bar.pack(fill=tk.X, padx=10, pady=5)

        self.progress_label = tk.Label(self, text="0%", font=(FONT_NAME, 10, "bold"), bg=bg_color, fg="white")
        self.progress_label.pack(pady=2)
        self.tree = ttk.Treeview(
            self,
            columns=['ID', 'Source', 'Entry Name', 'Path', 'Severity', 'Risk', 'Frequency', 'Activity Age', 'Timestamp',
                     'User', 'Exec Count', 'Signature'],
            show='headings',
            selectmode="extended"
        )
        cols = [
            ('ID', 50),
            ('Source', 150),
            ('Entry Name', 200),
            ('Path', 300),
            ('Severity', 100),
            ('Risk', 80),
            ('Frequency', 120),
            ('Activity Age', 120),
            ('Timestamp', 150),
            ('User', 120),
            ('Exec Count', 80),
            ('Signature', 200)
        ]
        for col, width in cols:
            self.tree.heading(col, text=col, command=lambda _col=col: self.sort_column(_col, False))
            self.tree.column(col, width=width, anchor=tk.W)
        style = ttk.Style()
        style.configure("Custom.Treeview.Heading", background="#333333", foreground="white", font=BOLD_FONT)
        self.tree.configure(style="Custom.Treeview")
        for sev, bg_color_tag in {"Normal": "#d4edda", "Suspicious": "#fff3cd", "Critical": "#f08080"}.items():
            self.tree.tag_configure(sev, background=bg_color_tag)
        self.tree.bind("<Double-1>", self.show_details)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.real_time_scanning = False
        self.scan_thread = None

    def add_suspicious_word(self):
        new_word = simpledialog.askstring("Add Suspicious Word", "Enter new suspicious word:")
        if new_word:
            if new_word.lower() in [w.lower() for w in DEFAULT_SUSPICIOUS_WORDS]:
                messagebox.showinfo("Info", f'The word "{new_word}" is already in the list.')
            else:
                DEFAULT_SUSPICIOUS_WORDS.append(new_word)
                messagebox.showinfo("Success", f'Word "{new_word}" added successfully.')
                self.refresh_dropdown()

    def delete_suspicious_word(self):
        word = self.selected_suspicious.get()
        if word == "All":
            messagebox.showinfo("Info", "Please select a specific word to delete.")

        elif word in DEFAULT_SUSPICIOUS_WORDS:
            DEFAULT_SUSPICIOUS_WORDS.remove(word)
            messagebox.showinfo("Success", f'Word "{word}" deleted.')
            self.refresh_dropdown()
        else:
            messagebox.showerror("Error", "Word not found in list.")

    def edit_suspicious_word(self):
        word = self.selected_suspicious.get()
        if word == "All":
            messagebox.showinfo("Info", "Please select a specific word to edit.")
        elif word in DEFAULT_SUSPICIOUS_WORDS:
            new_word = simpledialog.askstring("Edit Word", f"Enter new word for '{word}':")
            if new_word:
                index = DEFAULT_SUSPICIOUS_WORDS.index(word)
                DEFAULT_SUSPICIOUS_WORDS[index] = new_word
                messagebox.showinfo("Success", f'Word "{word}" changed to "{new_word}".')
                self.refresh_dropdown()
        else:
            messagebox.showerror("Error", "Word not found in list.")

    def refresh_dropdown(self):
        options = ["All"] + DEFAULT_SUSPICIOUS_WORDS
        menu = self.nametowidget(self.selected_suspicious._name + "_menu")
        menu.delete(0, "end")
        for option in options:
            menu.add_command(label=option, command=lambda value=option: self.selected_suspicious.set(value))

    def import_sigma_rules(self):
        # اختيار ملفات YAML لاستيراد قواعد SIGMA
        file_paths = filedialog.askopenfilenames(title="اختر ملف أو ملفات YAML لقواعد SIGMA",
                                                 filetypes=[("YAML Files", "*.yaml *.yml")])
        if not file_paths:
            return
        imported_count = import_sigma_rules_from_files(file_paths)
        messagebox.showinfo("Success", f"تم استيراد {imported_count} من قواعد SIGMA بنجاح.")
        self.controller.log_event(f"Imported {imported_count} SIGMA rules")

    def update_progress(self, value):
        self.progress_bar['value'] = value
        self.progress_label.config(text=f"{value}%")
        self.update_idletasks()

    def start_scan(self):
        self.controller.log_event("Scan started...")
        self.btn_start.config(state="disabled")
        if self.scan_type == "full":
            self.scan_thread = threading.Thread(
                target=lambda: [all_system_logs_scan_optimized(progress_callback=self.update_progress),
                                self.refresh_data_thread()])

            self.scan_thread.start()
        elif self.scan_type == "realtime":
            if not self.real_time_scanning:
                self.real_time_scanning = True
                self.scan_thread = threading.Thread(target=self.real_time_scan_loop, daemon=True)
                self.scan_thread.start()
                self.controller.log_event("Real-Time scanning activated")
            else:
                self.real_time_scanning = False
                self.controller.log_event("Real-Time scanning stopped")
        elif self.scan_type == "suspicious":
            self.scan_thread = threading.Thread(target=self.run_suspicious_scan)
            self.scan_thread.start()
        elif self.scan_type == "custom":
            self.scan_thread = threading.Thread(target=self.handle_custom_scan)
            self.scan_thread.start()
        self.after(100, lambda: self.btn_start.config(state="normal"))

    def stop_scan(self):
        if self.scan_type == "realtime" and self.real_time_scanning:
            self.real_time_scanning = False
            self.controller.log_event("Real-Time scan stopped")

    def real_time_scan_loop(self):
        while self.real_time_scanning:
            all_system_logs_scan_optimized(suppress_message=True, update_existing=True,
                                           progress_callback=self.update_progress)
            self.update_progress(0)
            self.refresh_data_thread()
            time.sleep(30)

    def run_suspicious_scan(self):
        word_filter = self.selected_suspicious.get() if self.scan_type == "suspicious" else "All"
        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT id, source, entry_name, path, severity, risk_percentage, frequency, activity_age, timestamp, current_user, execution_count, digital_signature
                    FROM registry_entries
                ''')
                all_records = cursor.fetchall()
                for rec in all_records:
                    rowid, source, entry_name, path, severity, risk, frequency, age, timestamp, current_user, exec_count, digital_sig = rec
                    if word_filter != "All":
                        if word_filter not in entry_name.lower() and word_filter not in path.lower():
                            continue
                    entry_dict = {
                        'severity': severity,
                        'risk_percentage': risk,

                        'execution_count': exec_count,
                        'digital_signature': digital_sig,
                        'path': path,
                        'hash': ''
                    }
                    new_sev = advanced_risk_assessment(entry_dict)
                    cursor.execute("UPDATE registry_entries SET severity=? WHERE id=?", (new_sev, rowid))
                conn.commit()
            messagebox.showinfo("Success", "Records re-evaluated based on suspicious word filter")
            self.controller.log_event("Suspicious scan completed and re-evaluated")
        except Exception as e:
            messagebox.showerror("Error", f"Scan failed: {str(e)}")
            log_exception("Suspicious scan error: " + str(e))
        finally:
            self.refresh_data_thread()

    def handle_custom_scan(self):
        self.controller.log_event("Custom scan started using Registry Tree")
        selected_entry = open_registry_tree_selection_dialog()
        if not selected_entry:
            self.controller.log_event("No registry entry selected")
            return
        try:
            current_user = getpass.getuser() if platform.system() == 'Windows' else os.getenv('USER', 'root')
        except Exception as e:
            log_exception("Error getting current user during custom scan: " + str(e))
            current_user = "Unknown"
        value = selected_entry.get("value", "")
        risk = calculate_risk_advanced(value, execution_count=1)
        entry = {
            'source': 'Registry Custom Scan',
            'entry_name': selected_entry.get("value_name", ""),
            'path': selected_entry.get("key_path", ""),
            'severity': 'Normal',
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'hash': hashlib.sha256(str(value).encode()).hexdigest(),
            'system_type': platform.system(),
            'current_user': current_user,
            'execution_count': 1,
            'digital_signature': hmac.new(HMAC_SECRET_KEY,
                                          hashlib.sha256(str(value).encode()).hexdigest().encode(),
                                          hashlib.sha256).hexdigest()[:12],
            'risk_percentage': risk
        }
        entry['severity'] = advanced_risk_assessment(entry)
        entry['frequency'] = compute_frequency(entry)
        entry['activity_age'] = compute_activity_age(entry)
        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                cursor.execute('''INSERT INTO registry_entries 

                                  (source, entry_name, path, severity, timestamp, hash, system_type,
                                   current_user, execution_count, digital_signature, risk_percentage, frequency, activity_age)
                                  VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                ''', (entry['source'], entry['entry_name'], entry['path'], entry['severity'],
                      entry['timestamp'], entry['hash'], entry['system_type'], entry['current_user'],
                      entry['execution_count'], entry['digital_signature'], entry['risk_percentage'],
                      entry['frequency'], entry['activity_age']))
                conn.commit()
            messagebox.showinfo("Success", f"Registry entry scanned: {entry['entry_name']}")
            self.controller.log_event("Custom scan completed successfully")
        except Exception as e:
            messagebox.showerror("Error", f"Scan failed: {str(e)}")
            log_exception("Custom scan error: " + str(e))
        finally:
            self.refresh_data_thread()

    def refresh_data_thread(self):
        threading.Thread(target=self.fetch_data, daemon=True).start()

    def fetch_data(self):
        try:
            if self.scan_type in ["realtime", "suspicious"]:
                time_window = "-5 minutes"
            else:
                time_window = "-30 minutes"
            query = f'''
                SELECT rowid as id, source, entry_name, path, severity, risk_percentage, frequency, activity_age, timestamp, current_user, execution_count, digital_signature 
                FROM registry_entries
                WHERE datetime(timestamp) >= datetime('now','{time_window}')
                ORDER BY id DESC
            '''
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.execute(query)
                rows = cursor.fetchall()
            self.after(0, lambda: self.update_tree(rows))
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Error", f"Error fetching data: {str(e)}"))
            log_exception("Data fetch error: " + str(e))

    def update_tree(self, rows):
        self.tree.delete(*self.tree.get_children())
        for i, row in enumerate(rows, start=1):
            new_row = (i,) + row[1:]
            self.tree.insert('', 'end', values=new_row, tags=(new_row[4],))
        self.controller.log_event("Data refreshed successfully")

    def delete_all_records(self):
        if not messagebox.askyesno("Confirmation", "Are you sure you want to delete all records?"):
            return

        try:
            with sqlite3.connect(DB_FILE) as conn:
                conn.execute("DELETE FROM registry_entries")
                conn.execute("UPDATE SQLITE_SEQUENCE SET SEQ=0 WHERE NAME='registry_entries'")
                conn.commit()
            self.refresh_data_thread()
            self.controller.log_event("All records deleted")
        except Exception as e:
            messagebox.showerror("Error", f"Deletion failed: {str(e)}")
            log_exception("Delete records error: " + str(e))

    def block_selected(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "No record selected")
            return
        for item in selected:
            self.block_entry(item)

    def block_entry(self, item_id):
        values = self.tree.item(item_id, "values")
        source = values[1]
        entry_name = values[2]
        path_info = values[3]
        if not messagebox.askyesno("Confirmation", f"Are you sure you want to block the record: {entry_name}?"):
            return
        if block_record(source, entry_name, path_info):
            with sqlite3.connect(DB_FILE) as conn:
                conn.execute("DELETE FROM registry_entries WHERE source=? AND entry_name=? AND path=?",
                             (source, entry_name, path_info))
                conn.commit()
            self.refresh_data_thread()
            self.controller.log_event(f"Record {entry_name} blocked successfully")

    def show_details(self, event):
        item_id = self.tree.focus()
        if not item_id:
            return
        values = self.tree.item(item_id, "values")
        detail_win = tk.Toplevel(self)
        detail_win.title("Entry Details")
        detail_text = ""
        columns = ['ID', 'Source', 'Entry Name', 'Path', 'Severity', 'Risk', 'Frequency', 'Activity Age', 'Timestamp',
                   'User', 'Exec Count', 'Signature']
        for col, val in zip(columns, values):
            detail_text += f"{col}: {val}\n"
        detail_label = tk.Label(detail_win, text=detail_text, font=(FONT_NAME, 10, "bold"), justify=tk.LEFT)
        detail_label.pack(padx=10, pady=10)

    def sort_column(self, col, reverse):

        l = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
        try:
            l.sort(key=lambda t: float(t[0]), reverse=reverse)
        except ValueError:
            l.sort(key=lambda t: t[0].lower(), reverse=reverse)
        for index, (val, k) in enumerate(l):
            self.tree.move(k, '', index)
        self.tree.heading(col, command=lambda: self.sort_column(col, not reverse))
        self.controller.log_event(f"Sorted column: {col}")

    def export_data(self):
        exp_win = tk.Toplevel(self)
        exp_win.title("Export Options")
        exp_win.grab_set()
        tk.Label(exp_win, text="Select export format:", font=BOLD_FONT).grid(row=0, column=0, padx=10, pady=5,
                                                                             sticky="w")
        formats = ['Excel', 'HTML', 'JSON', 'PDF']
        format_var = tk.StringVar(value='Excel')
        for idx, fmt in enumerate(formats):
            ttk.Radiobutton(exp_win, text=fmt, variable=format_var, value=fmt).grid(row=0, column=idx + 1, padx=5,
                                                                                    pady=5)
        tk.Label(exp_win, text="Select columns to export:", font=BOLD_FONT).grid(row=1, column=0, padx=10, pady=5,
                                                                                 sticky="w")
        columns = ['source', 'entry_name', 'path', 'severity', 'timestamp', 'current_user', 'execution_count',
                   'digital_signature']
        col_vars = {}
        for idx, col in enumerate(columns):
            var = tk.BooleanVar(value=True)
            col_vars[col] = var
            ttk.Checkbutton(exp_win, text=col, variable=var).grid(row=2 + idx // 3, column=idx % 3, padx=5, pady=2,
                                                                  sticky="w")

        def do_export():
            selected_cols = [col for col, var in col_vars.items() if var.get()]
            fmt = format_var.get()
            exp_win.destroy()
            if fmt == 'Excel':
                self.export_excel(selected_cols)
            elif fmt == 'HTML':
                self.export_html(selected_cols)
            elif fmt == 'JSON':
                self.export_json(selected_cols)
            elif fmt == 'PDF':
                self.export_pdf(selected_cols)

        ttk.Button(exp_win, text="Export", command=do_export).grid(row=6, column=1, padx=10, pady=10)

    def export_excel(self, columns):
        try:
            with sqlite3.connect(DB_FILE) as conn:
                df = pd.read_sql_query("SELECT {} FROM registry_entries".format(", ".join(columns)), conn)
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
            if file_path:

                details = pd.DataFrame({
                    'Scan Details': [f"Scan Type: Advanced Scan",
                                     f"Scan Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                })
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    details.to_excel(writer, index=False, header=False, startrow=0)
                    start_row = len(details) + 2
                    df.to_excel(writer, index=False, startrow=start_row)
                    worksheet = writer.sheets['Sheet1']
                    for idx, col in enumerate(df.columns, 1):
                        try:
                            col_letter = get_column_letter(idx)
                            max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                            worksheet.column_dimensions[col_letter].width = max_len
                        except Exception:
                            worksheet.column_dimensions[get_column_letter(idx)].width = 15
                messagebox.showinfo("Success", "Data exported to Excel successfully")
                self.controller.log_event("Exported to Excel: " + file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {str(e)}")
            self.controller.log_event("Excel export failed: " + str(e))

    def export_html(self, columns):
        try:
            with sqlite3.connect(DB_FILE) as conn:
                df = pd.read_sql_query("SELECT {} FROM registry_entries".format(", ".join(columns)), conn)
            file_path = filedialog.asksaveasfilename(defaultextension=".html")
            if file_path:
                details_html = f"""
                <h2>Scan Details</h2>
                <p>Scan Type: Advanced Scan<br>Scan Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                """
                style = f"""
                <style>
                table {{border-collapse: collapse; width: 100%;}}
                th, td {{border: 1px solid #ddd; padding: 8px; text-align: left;}}
                th {{background-color: #333333; color: white;}}
                </style>
                """
                html_table = df.to_html(index=False)
                html_data = f"<html><head><meta charset='utf-8'>{style}</head><body>{details_html}<br>{html_table}</body></html>"
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(html_data)
                messagebox.showinfo("Success", "Data exported to HTML successfully")
                self.controller.log_event("Exported to HTML: " + file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {str(e)}")
            self.controller.log_event("HTML export failed: " + str(e))

    def export_json(self, columns):
        try:
            with sqlite3.connect(DB_FILE) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.execute("SELECT {} FROM registry_entries".format(", ".join(columns)))
                data = [dict(row) for row in cursor.fetchall()]
            file_path = filedialog.asksaveasfilename(defaultextension=".json")
            if file_path:
                export_dict = {
                    "Scan Details": {
                        "Scan Type": "Advanced Scan",
                        "Scan Date": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    },
                    "Results": data
                }
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(export_dict, f, ensure_ascii=False, indent=4)
                messagebox.showinfo("Success", "Data exported to JSON successfully")
                self.controller.log_event("Exported to JSON: " + file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {str(e)}")
            self.controller.log_event("JSON export failed: " + str(e))

    def export_pdf(self, columns):
        try:
            with sqlite3.connect(DB_FILE) as conn:
                df = pd.read_sql_query("SELECT {} FROM registry_entries".format(", ".join(columns)), conn)
            file_path = filedialog.asksaveasfilename(defaultextension=".pdf")
            if not file_path:
                return
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            doc = SimpleDocTemplate(file_path, pagesize=landscape(A4))
            styles = getSampleStyleSheet()
            elements = []
            details_para = Paragraph(
                f"<b>Scan Details:</b> Advanced Scan, Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                styles['Normal']
            )
            elements.append(details_para)
            elements.append(Spacer(1, 12))
            data = [columns] + df.values.tolist()
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#333333")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONT', (0, 0), (-1, -1), 'Helvetica-Bold', 10),

                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ]))
            elements.append(table)
            doc.build(elements)
            messagebox.showinfo("Success", "Data exported to PDF successfully")
            self.controller.log_event("Exported to PDF: " + file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {str(e)}")
            self.controller.log_event("PDF export failed: " + str(e))


class SuspiciousScreen(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg=SUSPICIOUS_TAB_BG)
        self.controller = controller
        header = tk.Label(self, text="Suspicious Records", font=HEADER_FONT, bg=SUSPICIOUS_TAB_BG, fg="white")
        header.pack(pady=10)
        self.progress_bar = ttk.Progressbar(self, orient="horizontal", mode="determinate", maximum=100)
        self.progress_bar.pack(fill=tk.X, padx=10, pady=5)
        tool_frame = tk.Frame(self, bg=SUSPICIOUS_TAB_BG)
        tool_frame.pack(fill=tk.X, padx=10, pady=5)
        btn_refresh = tk.Button(tool_frame, text="Refresh Data", font=BOLD_FONT, bg=BUTTON_BG, fg="white",
                                command=self.refresh_data_thread, relief="raised", bd=3)
        btn_refresh.pack(side=tk.LEFT, padx=5)
        btn_export = tk.Button(tool_frame, text="Export Data", font=BOLD_FONT, bg=BUTTON_BG, fg="white",
                               command=self.export_data, relief="raised", bd=3)
        btn_export.pack(side=tk.LEFT, padx=5)
        btn_block = tk.Button(tool_frame, text="Block Record", font=BOLD_FONT, bg=BUTTON_BG, fg="white",
                              command=self.block_system_entry, relief="raised", bd=3)
        btn_block.pack(side=tk.LEFT, padx=5)
        self.tree = ttk.Treeview(
            self,
            columns=['ID', 'Source', 'Entry Name', 'Path', 'Severity', 'Risk', 'Frequency', 'Activity Age', 'Timestamp',
                     'User', 'Exec Count', 'Signature'],
            show='headings',
            selectmode="extended"
        )
        cols = [
            ('ID', 50),
            ('Source', 150),
            ('Entry Name', 200),
            ('Path', 300),
            ('Severity', 100),
            ('Risk', 80),
            ('Frequency', 120),
            ('Activity Age', 120),
            ('Timestamp', 150),
            ('User', 120),
            ('Exec Count', 80),
            ('Signature', 200)

        ]
        for col, width in cols:
            self.tree.heading(col, text=col, command=lambda _col=col: self.sort_column(_col, False))
            self.tree.column(col, width=width, anchor=tk.W)
        style = ttk.Style()
        style.configure("Custom.Treeview.Heading", background="#333333", foreground="white", font=BOLD_FONT)
        self.tree.configure(style="Custom.Treeview")
        for sev, bg_color in {"Normal": "#d4edda", "Suspicious": "#fff3cd", "Critical": "#f08080"}.items():
            self.tree.tag_configure(sev, background=bg_color)
        self.tree.bind("<Double-1>", self.show_details)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.refresh_data_thread()

    def refresh_data_thread(self):
        threading.Thread(target=self.fetch_data, daemon=True).start()

    def fetch_data(self):
        try:
            query = '''
                SELECT rowid as id, source, entry_name, path, severity, risk_percentage, frequency, activity_age, timestamp, current_user, execution_count, digital_signature 
                FROM registry_entries
                WHERE severity IN ('Suspicious','Critical')
                  AND datetime(timestamp) >= datetime('now','-5 minutes')
                ORDER BY id DESC
            '''
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.execute(query)
                rows = cursor.fetchall()
            self.after(0, lambda: self.update_tree(rows))
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Error", f"Error fetching data: {str(e)}"))
            log_exception("Suspicious data fetch error: " + str(e))

    def update_tree(self, rows):
        self.tree.delete(*self.tree.get_children())
        for i, row in enumerate(rows, start=1):
            new_row = (i,) + row[1:]
            self.tree.insert('', 'end', values=new_row, tags=(new_row[4],))
        self.controller.log_event("Data refreshed successfully")

    def sort_column(self, col, reverse):
        l = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
        try:
            l.sort(key=lambda t: float(t[0]), reverse=reverse)
        except ValueError:
            l.sort(key=lambda t: t[0].lower(), reverse=reverse)
        for index, (val, k) in enumerate(l):
            self.tree.move(k, '', index)
        self.tree.heading(col, command=lambda: self.sort_column(col, not reverse))

        self.controller.log_event(f"Sorted column: {col}")

    def show_details(self, event):
        item_id = self.tree.focus()
        if not item_id:
            return
        values = self.tree.item(item_id, "values")
        detail_win = tk.Toplevel(self)
        detail_win.title("Entry Details")
        detail_text = ""
        columns = ['ID', 'Source', 'Entry Name', 'Path', 'Severity', 'Risk', 'Frequency', 'Activity Age', 'Timestamp',
                   'User', 'Exec Count', 'Signature']
        for col, val in zip(columns, values):
            detail_text += f"{col}: {val}\n"
        detail_label = tk.Label(detail_win, text=detail_text, font=(FONT_NAME, 10, "bold"), justify=tk.LEFT)
        detail_label.pack(padx=10, pady=10)

    def block_system_entry(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "No record selected")
            return
        for item in selected:
            self.block_entry(item)

    def block_entry(self, item_id):
        values = self.tree.item(item_id, "values")
        source = values[1]
        entry_name = values[2]
        path_info = values[3]
        if not messagebox.askyesno("Confirmation", f"Are you sure you want to block the record: {entry_name}?"):
            return
        if block_record(source, entry_name, path_info):
            with sqlite3.connect(DB_FILE) as conn:
                conn.execute("DELETE FROM registry_entries WHERE source=? AND entry_name=? AND path=?",
                             (source, entry_name, path_info))
                conn.commit()
            self.refresh_data_thread()
            self.controller.log_event(f"Record {entry_name} blocked successfully")

    def export_data(self):
        exp_win = tk.Toplevel(self)
        exp_win.title("Export Options")
        exp_win.grab_set()
        tk.Label(exp_win, text="Select export format:", font=BOLD_FONT).grid(row=0, column=0, padx=10, pady=5,
                                                                             sticky="w")
        formats = ['Excel', 'HTML', 'JSON', 'PDF']
        format_var = tk.StringVar(value='Excel')
        for idx, fmt in enumerate(formats):
            ttk.Radiobutton(exp_win, text=fmt, variable=format_var, value=fmt).grid(row=0, column=idx + 1,

                                                                                    padx=5, pady=5)
        tk.Label(exp_win, text="Select columns to export:", font=BOLD_FONT).grid(row=1, column=0, padx=10, pady=5,
                                                                                 sticky="w")
        columns = ['source', 'entry_name', 'path', 'severity', 'timestamp', 'current_user', 'execution_count',
                   'digital_signature']
        col_vars = {}
        for idx, col in enumerate(columns):
            var = tk.BooleanVar(value=True)
            col_vars[col] = var
            ttk.Checkbutton(exp_win, text=col, variable=var).grid(row=2 + idx // 3, column=idx % 3, padx=5, pady=2,
                                                                  sticky="w")

        def do_export():
            selected_cols = [col for col, var in col_vars.items() if var.get()]
            fmt = format_var.get()
            exp_win.destroy()
            if fmt == 'Excel':
                self.export_excel(selected_cols)
            elif fmt == 'HTML':
                self.export_html(selected_cols)
            elif fmt == 'JSON':
                self.export_json(selected_cols)
            elif fmt == 'PDF':
                self.export_pdf(selected_cols)

        ttk.Button(exp_win, text="Export", command=do_export).grid(row=6, column=1, padx=10, pady=10)


if __name__ == "__main__":
    app = MainApp()
    app.mainloop()









































